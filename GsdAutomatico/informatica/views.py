# GsdAutomatico/informatica/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.apps import apps
from django.urls import reverse, NoReverseMatch, reverse_lazy
from django.contrib.auth.models import User, Group
from Ouvidoria.models import PATD, Anexo, Configuracao, AlegacaoDefesaLog
from Secao_pessoal.models import Efetivo, Setor # ATUALIZADO
from login.models import UserProfile
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .forms import (
    InformaticaUserCreationForm, InformaticaUserChangeForm,
    GroupForm, ConfiguracaoForm, BackupDestinoForm
)
from .models import (
    GrupoMaterial, SubgrupoMaterial, Material, Cautela, CautelaItem, Armario, Prateleira,
    GroupProfile, SECAO_CHOICES, ConfiguracaoComandantes, BackupDestino, BackupExecucao,
)
from django.db import transaction
from django.db.models import Q, ProtectedError
import logging
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils import timezone
from django.core.cache import cache
import json
import datetime
import docker
import os
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate # Importação para validar senha
from auditoria.utils import registrar, resolver_label

logger = logging.getLogger(__name__)

_INFORMATICA_PERMISSAO_MAP = {
    'informatica-admin': 'Admin- Informática',
    'informatica-secao': 'Seção- Informática',
}

def is_informatica_admin(user):
    return user.is_superuser or user.groups.filter(name='informatica-admin').exists()

def is_informatica_secao(user):
    return user.is_superuser or user.groups.filter(name__in=['informatica-admin', 'informatica-secao']).exists()

class InformaticaAdminMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self): return is_informatica_admin(self.request.user)
    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            from django.shortcuts import redirect
            return redirect('home:index')
        return super().handle_no_permission()

class InformaticaAcessoMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self): return is_informatica_secao(self.request.user)
    def handle_no_permission(self): return super().handle_no_permission()

# manter StaffRequiredMixin como alias para não quebrar código existente
class StaffRequiredMixin(InformaticaAdminMixin):
    pass

def is_staff(user):
    return is_informatica_admin(user)

# ==========================================
# VIEWS DE DASHBOARD E CRUD BÁSICO
# ==========================================
@login_required
def dashboard(request):
    if not is_informatica_admin(request.user):
        return redirect('informatica:gestao_materiais')
    general_models_info = [
        ('auth', 'user', 'Utilizadores', 'Utilizador'),
        ('auth', 'group', 'Grupos de Permissão', 'Grupo'),
    ]
    ouvidoria_models_info = [
        ('Secao_pessoal', 'efetivo', 'Efetivo (Militares)', 'Militar'),
        ('Ouvidoria', 'patd', 'Processos (PATD)', 'PATD'),
        ('Ouvidoria', 'configuracao', 'Configurações Gerais', 'Configuração'),
    ]

    def get_model_admin_links(model_info_list):
        processed_list = []
        for app_label, model_name, plural_name, singular_name in model_info_list:
            item_data = {
                'name': plural_name, 'singular': singular_name, 'model_name': model_name,
                'list_url': '#', 'add_url': '#', 'edit_config_url': '#', 'count': None
            }
            try:
                model = apps.get_model(app_label, model_name)
                list_url_name = f'informatica:{model_name}_list'
                add_url_name = f'informatica:{model_name}_add'
                edit_config_url_name = f'informatica:configuracao_edit'

                if model_name == 'configuracao':
                    try:
                        item_data['edit_config_url'] = reverse(edit_config_url_name)
                        item_data['list_url'] = item_data['edit_config_url']
                        item_data['add_url'] = item_data['edit_config_url']
                    except NoReverseMatch: pass
                else:
                    try: item_data['list_url'] = reverse(list_url_name)
                    except NoReverseMatch: pass
                    try: item_data['add_url'] = reverse(add_url_name)
                    except NoReverseMatch: pass
                    item_data['count'] = model.objects.count()

                processed_list.append(item_data)
            except Exception as e: logger.error(f"Erro: {e}")
        return processed_list

    general_admin_apps = get_model_admin_links(general_models_info)
    ouvidoria_apps = get_model_admin_links(ouvidoria_models_info)
    
    quick_stats = cache.get('informatica_dashboard_stats')
    if quick_stats is None:
        quick_stats = {
            'total_users': User.objects.count(),
            'acervo': Material.objects.count(),
            'em_cautela': Cautela.objects.filter(ativa=True).count(),
            'militares_count': Efetivo.objects.count(),
        }
        cache.set('informatica_dashboard_stats', quick_stats, timeout=300)

    terminal_logs = cache.get('docker_terminal_logs', [
        f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Atualizando logs em background..."
    ])

    cautelas_recentes = Cautela.objects.filter(ativa=True).select_related('recebedor').prefetch_related('itens__material').order_by('-data_emissao')[:5]

    from caixa_entrada.models import Mensagem
    from chamados.models import Chamado
    user = request.user
    count_inbox = (
        Mensagem.objects
        .filter(destinatarios=user, eh_rascunho=False)
        .exclude(excluida_por=user)
        .exclude(lida_por=user)
        .count()
    )
    count_chamados = Chamado.objects.filter(
        atribuido_a__isnull=True
    ).exclude(status__in=['resolvido', 'fechado']).count()

    context = {
        'page_title': 'Dashboard da Informática',
        'general_admin_apps': general_admin_apps,
        'ouvidoria_apps': ouvidoria_apps,
        'quick_stats': quick_stats,
        'terminal_logs': terminal_logs,
        'cautelas_recentes': cautelas_recentes,
        'count_inbox': count_inbox,
        'count_chamados': count_chamados,
    }
    return render(request, 'informatica/dashboard.html', context)


class UserListView(StaffRequiredMixin, ListView):
    model = User
    template_name = 'informatica/user_list.html'
    context_object_name = 'users'
    paginate_by = 20
    def get_queryset(self):
        queryset = super().get_queryset().select_related('profile__militar').order_by('username')
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(username__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query) |
                Q(profile__militar__nome_guerra__icontains=query)
            ).distinct()
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Gerenciar Utilizadores"
        context['search_query'] = self.request.GET.get('q', '')
        return context

def _get_grupos_por_secao():
    """Grupos organizados por seção, usando GroupProfile.secao."""
    from collections import defaultdict
    grupos = Group.objects.select_related('secao_profile').order_by('name')
    secao_labels = dict(SECAO_CHOICES)
    por_secao = defaultdict(list)
    for g in grupos:
        profile = getattr(g, 'secao_profile', None)
        key = profile.secao if profile and profile.secao in secao_labels else 'geral'
        por_secao[key].append(g)
    return [
        {'key': key, 'label': label, 'groups': por_secao[key]}
        for key, label in SECAO_CHOICES
        if por_secao.get(key)
    ]


class UserCreateView(StaffRequiredMixin, CreateView):
    model = User
    form_class = InformaticaUserCreationForm
    template_name = 'informatica/user_form.html'
    success_url = reverse_lazy('informatica:user_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Adicionar Utilizador"
        context['grupos_por_secao'] = _get_grupos_por_secao()
        context['selected_group_ids'] = set()
        return context
    def form_valid(self, form):
        user = form.save()
        generated_pwd = getattr(user, '_generated_password', '—')
        registrar(
            self.request.user, secao='informatica',
            permissao=resolver_label(self.request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='criou', descricao=f"criou o usuário '{user.username}'",
            objeto_tipo='Usuário', objeto_id=user.username,
        )
        messages.success(self.request, f"Utilizador '{user.username}' criado. Senha temporária: {generated_pwd} (anote agora — não será exibida novamente).")
        return redirect(self.success_url)

class UserUpdateView(StaffRequiredMixin, UpdateView):
    model = User
    form_class = InformaticaUserChangeForm
    template_name = 'informatica/user_form.html'
    success_url = reverse_lazy('informatica:user_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f"Editar Utilizador: {self.object.username}"
        context['grupos_por_secao'] = _get_grupos_por_secao()
        context['selected_group_ids'] = set(self.object.groups.values_list('pk', flat=True))
        try:
            context['militar_atual'] = self.object.profile.militar
        except Exception:
            context['militar_atual'] = None
        return context
    def form_valid(self, form):
        response = super().form_valid(form)
        registrar(
            self.request.user, secao='informatica',
            permissao=resolver_label(self.request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='editou', descricao=f"editou o usuário '{self.object.username}'",
            objeto_tipo='Usuário', objeto_id=self.object.username,
        )
        return response

class UserDeleteView(StaffRequiredMixin, DeleteView):
    model = User
    template_name = 'informatica/user_confirm_delete.html'
    success_url = reverse_lazy('informatica:user_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f"Confirmar Exclusão de {self.object.username}"
        return context
    def form_valid(self, form):
        username = self.object.username
        registrar(
            self.request.user, secao='informatica',
            permissao=resolver_label(self.request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='excluiu', descricao=f"excluiu o usuário '{username}'",
            objeto_tipo='Usuário', objeto_id=username,
        )
        return super().form_valid(form)

@staff_member_required
@require_POST
def reset_user_password(request, pk):
    user_to_reset = get_object_or_404(User, pk=pk)
    try:
        temp_password = '12345678'
        user_to_reset.set_password(temp_password)
        user_to_reset.save()
        registrar(
            request.user, secao='informatica',
            permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='editou', descricao=f"redefiniu a senha do usuário '{user_to_reset.username}'",
            objeto_tipo='Usuário', objeto_id=user_to_reset.username,
        )
        return JsonResponse({
            'status': 'success',
            'message': f"Senha do utilizador '{user_to_reset.username}' redefinida para 12345678.",
            'nova_senha': temp_password,
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': 'Não foi possível redefinir a senha.'}, status=500)

class GroupListView(StaffRequiredMixin, ListView):
    model = Group
    template_name = 'informatica/group_list.html'
    context_object_name = 'groups'
    paginate_by = 9999  # agrupamos manualmente, sem paginação por página
    def get_queryset(self):
        queryset = super().get_queryset().select_related('secao_profile').order_by('name')
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(name__icontains=query)
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Gerenciar Grupos"
        context['search_query'] = self.request.GET.get('q', '')
        context['secao_ativa'] = self.request.GET.get('secao', 'todas')
        context['secao_choices'] = SECAO_CHOICES
        # Agrupa grupos por seção
        from collections import defaultdict
        grupos_por_secao = defaultdict(list)
        for g in context['groups']:
            secao = getattr(g.secao_profile, 'secao', 'geral') if hasattr(g, 'secao_profile') else 'geral'
            grupos_por_secao[secao].append(g)
        # Garante que todas as seções com grupos apareçam, na ordem de SECAO_CHOICES
        secao_labels = dict(SECAO_CHOICES)
        context['grupos_por_secao'] = [
            {'key': key, 'label': label, 'groups': grupos_por_secao.get(key, [])}
            for key, label in SECAO_CHOICES
            if grupos_por_secao.get(key)
        ]
        return context

def _get_perms_por_app():
    from django.contrib.auth.models import Permission
    from collections import defaultdict
    agrupado = defaultdict(lambda: {'label': '', 'perms': []})
    for perm in Permission.objects.select_related('content_type').order_by(
            'content_type__app_label', 'content_type__model', 'codename'):
        app = perm.content_type.app_label
        agrupado[app]['label'] = app.replace('_', ' ').title()
        agrupado[app]['perms'].append({'id': perm.pk, 'name': perm.name, 'codename': perm.codename})
    return dict(agrupado)

class GroupCreateView(StaffRequiredMixin, CreateView):
    model = Group
    form_class = GroupForm
    template_name = 'informatica/group_form.html'
    success_url = reverse_lazy('informatica:group_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        secao = self.request.POST.get('secao', 'geral')
        GroupProfile.objects.update_or_create(group=self.object, defaults={'secao': secao})
        registrar(
            self.request.user, secao='informatica',
            permissao=resolver_label(self.request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='criou', descricao=f"criou o grupo '{self.object.name}'",
            objeto_tipo='Grupo', objeto_id=self.object.name,
        )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Novo Grupo"
        context['perms_por_app'] = _get_perms_por_app()
        context['selected_perm_ids'] = []
        context['secao_choices'] = SECAO_CHOICES
        context['secao_atual'] = 'geral'
        return context

class GroupUpdateView(StaffRequiredMixin, UpdateView):
    model = Group
    form_class = GroupForm
    template_name = 'informatica/group_form.html'
    success_url = reverse_lazy('informatica:group_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        secao = self.request.POST.get('secao', 'geral')
        GroupProfile.objects.update_or_create(group=self.object, defaults={'secao': secao})
        registrar(
            self.request.user, secao='informatica',
            permissao=resolver_label(self.request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='editou', descricao=f"editou o grupo '{self.object.name}'",
            objeto_tipo='Grupo', objeto_id=self.object.name,
        )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f"Editar Grupo: {self.object.name}"
        context['perms_por_app'] = _get_perms_por_app()
        context['selected_perm_ids'] = list(self.object.permissions.values_list('id', flat=True))
        context['secao_choices'] = SECAO_CHOICES
        profile = GroupProfile.objects.filter(group=self.object).first()
        context['secao_atual'] = profile.secao if profile else 'geral'
        return context

class GroupDeleteView(StaffRequiredMixin, DeleteView):
    model = Group
    template_name = 'informatica/group_confirm_delete.html'
    success_url = reverse_lazy('informatica:group_list')

    def form_valid(self, form):
        nome = self.object.name
        registrar(
            self.request.user, secao='informatica',
            permissao=resolver_label(self.request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='excluiu', descricao=f"excluiu o grupo '{nome}'",
            objeto_tipo='Grupo', objeto_id=nome,
        )
        GroupProfile.objects.filter(group=self.object).delete()
        return super().form_valid(form)


class ConfiguracaoUpdateView(StaffRequiredMixin, UpdateView):
    model = Configuracao
    form_class = ConfiguracaoForm
    template_name = 'informatica/configuracao_form.html'
    success_url = reverse_lazy('informatica:configuracao_edit')
    def get_object(self, queryset=None):
        obj, created = Configuracao.objects.get_or_create(pk=1)
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from django.utils import timezone
        from datetime import timedelta
        config = Configuracao.load()
        cutoff = timezone.now() - timedelta(days=config.dias_retencao_lixeira)
        lixeira_patds = PATD.all_objects.filter(deleted=True).select_related('militar').order_by('deleted_at')
        lixeira_list = []
        for p in lixeira_patds:
            if p.deleted_at:
                expira_em = p.deleted_at + timedelta(days=config.dias_retencao_lixeira)
                dias_restantes = (expira_em - timezone.now()).days
            else:
                expira_em = None
                dias_restantes = None
            lixeira_list.append({
                'id': p.id,
                'numero_patd': p.numero_patd,
                'militar': str(p.militar),
                'deleted_at': p.deleted_at,
                'expira_em': expira_em,
                'dias_restantes': dias_restantes,
            })
        ctx['lixeira_list'] = lixeira_list
        ctx['lixeira_total'] = len(lixeira_list)
        ctx['dias_retencao_lixeira'] = config.dias_retencao_lixeira
        return ctx


# ==========================================
# COMANDANTES
# ==========================================
@staff_member_required
def configuracao_comandantes(request):
    config = ConfiguracaoComandantes.get_instance()
    config_ouvidoria = Configuracao.load()
    oficiais = Efetivo.objects.filter(oficial=True).order_by('posto', 'nome_guerra')
    if request.method == 'POST':
        def _get(field):
            pk = request.POST.get(field)
            return Efetivo.objects.filter(pk=pk).first() if pk else None
        config.comandante_gsd = _get('comandante_gsd')
        config.comandante_bagl = _get('comandante_bagl')
        config.chefe_sop = _get('chefe_sop')
        config.comandante_esi = _get('comandante_esi')
        config.save()
        config_ouvidoria.oficial_chefe_ouvidoria = _get('oficial_chefe_ouvidoria')
        config_ouvidoria.save()
        registrar(
            request.user, secao='informatica',
            permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='editou', descricao="editou a configuração de comandantes",
            objeto_tipo='ConfiguracaoComandantes', objeto_id='1',
        )
        from django.contrib import messages
        messages.success(request, 'Comandantes salvos com sucesso.')
        return redirect('informatica:configuracao_comandantes')
    return render(request, 'informatica/configuracao_comandantes.html', {
        'config': config,
        'config_ouvidoria': config_ouvidoria,
        'oficiais': oficiais,
    })


# ==========================================
# LOGS E MONITORAMENTO
# ==========================================
@staff_member_required
def system_logs_api(request):
    from .tasks import fetch_docker_logs_task
    logs_data = cache.get('docker_logs_api')
    if logs_data is None:
        fetch_docker_logs_task.delay()
        logs_data = [{'container': 'system', 'text': 'Coletando logs em background, tente novamente em instantes...'}]
    return JsonResponse({'logs': logs_data})

# ==========================================
# BACKUP (Fase 2) — substitui o antigo monitoramento_backup/visualizar_logs_backup
# ==========================================
class BackupDestinoUpdateView(StaffRequiredMixin, UpdateView):
    model = BackupDestino
    form_class = BackupDestinoForm
    template_name = 'informatica/backup_destino_form.html'
    success_url = reverse_lazy('informatica:backup_destino_config')

    def get_object(self, queryset=None):
        return BackupDestino.get_instance()

    def form_valid(self, form):
        messages.success(self.request, 'Configuração do servidor reserva salva com sucesso.')
        return super().form_valid(form)


class BackupHistoricoListView(InformaticaAdminMixin, ListView):
    model = BackupExecucao
    template_name = 'informatica/backup_historico.html'
    context_object_name = 'execucoes'
    paginate_by = 30


@login_required
@require_POST
def backup_executar_agora(request):
    if not is_informatica_admin(request.user):
        return JsonResponse({'erro': 'Acesso negado.'}, status=403)
    from .tasks import executar_backup_manual_task
    from .models import BackupExecucao
    execucao = BackupExecucao.objects.create()
    executar_backup_manual_task.apply_async(kwargs={'execucao_id': execucao.pk})
    return JsonResponse({'execucao_id': execucao.pk})


@login_required
def backup_status_json(request):
    if not is_informatica_admin(request.user):
        return JsonResponse({'erro': 'Acesso negado.'}, status=403)
    from .models import BackupExecucao
    try:
        execucao = BackupExecucao.objects.get(pk=request.GET.get('id'))
    except (BackupExecucao.DoesNotExist, ValueError, TypeError):
        return JsonResponse({'erro': 'Não encontrado.'}, status=404)
    return JsonResponse({
        'status': execucao.status,
        'erro_detalhe': execucao.erro_detalhe or '',
    })


@login_required
def backup_explorar(request, pk):
    """Busca um registro específico (PATD, Efetivo, etc.) dentro de um backup antigo e
    compara campo a campo com o registro atual — sem alterar nada até confirmar a restauração."""
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    from .backup_diff import (MODELOS_DIFF, PATD_STATUS_LABELS,
                               restaurar_dump_temp, dropar_temp,
                               buscar_registro_temp, montar_diff,
                               buscar_registro_atual, listar_todos_temp,
                               listar_ausentes_no_sistema)

    execucao = get_object_or_404(BackupExecucao, pk=pk)
    modelo_key = request.GET.get('modelo', '')
    valor      = request.GET.get('valor', '').strip()
    filtro     = request.GET.get('filtro', '').strip().lower()
    modo_ausentes = request.GET.get('ausentes') == '1'
    diffs = None
    registro_pk = None
    old_dict_destaque = None
    erro = None
    lista_registros = None
    colunas_lista = []
    config = MODELOS_DIFF.get(modelo_key)
    arquivo_local = None
    arquivo_temp = None

    if config:
        model = config['model']
        colunas_lista = config.get('colunas_lista', [])

        arquivo_local, arquivo_temp = _resolver_arquivo_db(execucao)
        if not arquivo_local:
            erro = 'Arquivo de backup não encontrado em disco nem no servidor remoto.'
        elif valor:
            # ── Modo detalhe: busca individual e diff ──
            live_obj = buscar_registro_atual(model, config['busca_campo'], valor)
            if not live_obj:
                # tenta apenas pelo backup sem registro atual (PATD deletada)
                tempdb = None
                try:
                    tempdb = restaurar_dump_temp(arquivo_local)
                    pk_int = int(valor) if valor.isdigit() else None
                    old_dict_destaque = buscar_registro_temp(tempdb, model, pk_int) if pk_int else None
                    if old_dict_destaque is None:
                        erro = 'Registro não encontrado no banco atual nem no backup.'
                    else:
                        # Sem live_obj: exibe apenas dados do backup sem diff
                        diffs = [
                            {'campo': k, 'label': k, 'antigo': v, 'atual': '—', 'diferente': False}
                            for k, v in old_dict_destaque.items()
                            if k not in ('senha_unica', 'senha_criptografada', 'password')
                        ]
                        registro_pk = pk_int
                except Exception as exc:
                    erro = f'Erro ao ler o backup: {exc}'
                finally:
                    if tempdb:
                        dropar_temp(tempdb)
            else:
                tempdb = None
                try:
                    tempdb = restaurar_dump_temp(arquivo_local)
                    old_dict = buscar_registro_temp(tempdb, model, live_obj.pk)
                    if old_dict is None:
                        erro = 'Este registro ainda não existia nesse backup.'
                    else:
                        diffs = montar_diff(old_dict, live_obj)
                        old_dict_destaque = old_dict
                        registro_pk = live_obj.pk
                except Exception as exc:
                    erro = f'Erro ao restaurar/ler o backup: {exc}'
                finally:
                    if tempdb:
                        dropar_temp(tempdb)
        else:
            # ── Modo listagem (todos ou apenas ausentes no sistema) ──
            tempdb = None
            try:
                tempdb = restaurar_dump_temp(arquivo_local)
                if modo_ausentes:
                    raw = listar_ausentes_no_sistema(tempdb, model)
                else:
                    raw = listar_todos_temp(tempdb, model)

                def _cell(row, col, tipo):
                    v = row.get(col)
                    if tipo == 'date' and v:
                        return str(v)[:10]
                    if tipo == 'bool_sim_nao':
                        return 'Sim' if v else 'Não'
                    if tipo == 'bool_lixeira':
                        return '🗑 Sim' if v else '—'
                    if tipo == 'status_patd':
                        return PATD_STATUS_LABELS.get(v, v or '—')
                    return str(v) if v is not None else '—'

                def _badge(row):
                    """Retorna (texto, classe_css) para o badge de situação."""
                    if row.get('deleted'):
                        return ('Lixeira',   'danger')
                    if row.get('arquivado'):
                        return ('Arquivado', 'secondary')
                    st = row.get('status', '')
                    if st == 'finalizado':
                        return ('Finalizado','success')
                    if st:
                        return (PATD_STATUS_LABELS.get(st, st), 'info')
                    if row.get('ativo') is False or row.get('ativa') is False:
                        return ('Inativo', 'secondary')
                    return (None, None)

                registros = []
                for r in raw:
                    cells = [_cell(r, col, tipo) for col, _, tipo in colunas_lista]
                    badge_txt, badge_cls = _badge(r)
                    row_text = ' '.join(str(c) for c in cells).lower()
                    if filtro and filtro not in row_text:
                        continue
                    registros.append({
                        'id': r.get('id'),
                        'cells': cells,
                        'badge_txt': badge_txt,
                        'badge_cls': badge_cls,
                    })
                lista_registros = registros
            except Exception as exc:
                erro = f'Erro ao ler o backup: {exc}'
            finally:
                if tempdb:
                    dropar_temp(tempdb)

    # Limpa arquivo temporário baixado do servidor remoto (se houver)
    if arquivo_temp and os.path.exists(arquivo_temp):
        try:
            os.remove(arquivo_temp)
        except Exception:
            pass

    # Campos de destaque para a view de detalhe
    campos_destaque = []
    if old_dict_destaque and config:
        for campo in config.get('campos_destaque', []):
            v = old_dict_destaque.get(campo)
            if v:
                campos_destaque.append({'campo': campo, 'valor': v})

    return render(request, 'informatica/backup_explorar.html', {
        'execucao': execucao,
        'modelos': MODELOS_DIFF,
        'modelo_key': modelo_key,
        'valor': valor,
        'filtro': filtro,
        'modo_ausentes': modo_ausentes,
        'diffs': diffs,
        'registro_pk': registro_pk,
        'erro': erro,
        'lista_registros': lista_registros,
        'colunas_lista': colunas_lista,
        'campos_destaque': campos_destaque,
    })


@login_required
@require_POST
def backup_restaurar_registro(request, pk):
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    from .backup_diff import MODELOS_DIFF, restaurar_dump_temp, dropar_temp, buscar_registro_temp, aplicar_restore

    execucao = get_object_or_404(BackupExecucao, pk=pk)
    modelo_key = request.POST.get('modelo', '')
    valor = request.POST.get('valor', '')
    registro_pk = request.POST.get('registro_pk')
    campos_selecionados = request.POST.getlist('campos')
    redirect_url = f"{reverse('informatica:backup_explorar', args=[pk])}?modelo={modelo_key}&valor={valor}"

    if modelo_key not in MODELOS_DIFF or not registro_pk or not campos_selecionados:
        messages.error(request, 'Selecione ao menos um campo para restaurar.')
        return redirect(redirect_url)

    model = MODELOS_DIFF[modelo_key]['model']
    live_obj = get_object_or_404(model, pk=registro_pk)

    arquivo_local, arquivo_temp = _resolver_arquivo_db(execucao)
    if not arquivo_local:
        messages.error(request, 'Arquivo de backup não encontrado em disco nem no servidor remoto.')
        return redirect(redirect_url)

    tempdb = None
    try:
        tempdb = restaurar_dump_temp(arquivo_local)
        old_dict = buscar_registro_temp(tempdb, model, live_obj.pk)
        if old_dict is None:
            messages.error(request, 'Registro não encontrado no backup.')
        else:
            alterados = aplicar_restore(live_obj, old_dict, campos_selecionados)
            if alterados:
                logger.info(
                    "[BACKUP RESTORE] %s restaurou campos %s do registro %s (%s) a partir do backup #%s",
                    request.user.username, alterados, model.__name__, live_obj.pk, execucao.pk,
                )
                registrar(
                    request.user, secao='informatica',
                    permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
                    acao='restaurou',
                    descricao=f"restaurou campos ({', '.join(alterados)}) do {model.__name__} #{live_obj.pk} via backup #{execucao.pk}",
                    objeto_tipo=model.__name__, objeto_id=str(live_obj.pk),
                )
                messages.success(request, f"Campos restaurados: {', '.join(alterados)}.")
            else:
                messages.info(request, 'Nenhuma alteração necessária — os valores já eram iguais.')
    except Exception as exc:
        messages.error(request, f'Erro ao restaurar: {exc}')
    finally:
        if tempdb:
            dropar_temp(tempdb)
        if arquivo_temp and os.path.exists(arquivo_temp):
            try:
                os.remove(arquivo_temp)
            except Exception:
                pass

    return redirect(redirect_url)


@login_required
@require_POST
def backup_restaurar_em_lote(request, pk):
    """Restaura múltiplos registros selecionados na listagem do backup.

    Para cada ID selecionado:
    - Se o registro já existe no banco (ativo ou soft-deleted): atualiza todos os campos.
    - Se não existe de jeito nenhum: recria com o PK original do backup.
    """
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    from .backup_diff import (MODELOS_DIFF, restaurar_dump_temp, dropar_temp,
                               buscar_registro_temp, aplicar_restore,
                               campos_comparaveis, recriar_registro)

    execucao = get_object_or_404(BackupExecucao, pk=pk)
    modelo_key = request.POST.get('modelo', '')
    modo_ausentes = request.POST.get('ausentes', '0')
    pks_str = request.POST.getlist('pks')
    redirect_url = (
        f"{reverse('informatica:backup_explorar', args=[pk])}"
        f"?modelo={modelo_key}"
        + ('&ausentes=1' if modo_ausentes == '1' else '')
    )

    if modelo_key not in MODELOS_DIFF or not pks_str:
        messages.error(request, 'Selecione ao menos um registro para restaurar.')
        return redirect(redirect_url)

    try:
        pks_int = [int(p) for p in pks_str]
    except ValueError:
        messages.error(request, 'IDs inválidos.')
        return redirect(redirect_url)

    model = MODELOS_DIFF[modelo_key]['model']
    manager = getattr(model, 'all_objects', model.objects)
    todos_campos = [f.name for f in campos_comparaveis(model)]

    arquivo_local, arquivo_temp = _resolver_arquivo_db(execucao)
    if not arquivo_local:
        messages.error(request, 'Arquivo de backup não encontrado em disco nem no servidor remoto.')
        return redirect(redirect_url)

    criados = 0
    atualizados = 0
    erros = []
    tempdb = None
    try:
        tempdb = restaurar_dump_temp(arquivo_local)
        for reg_pk in pks_int:
            old_dict = buscar_registro_temp(tempdb, model, reg_pk)
            if old_dict is None:
                erros.append(f'ID {reg_pk}: não encontrado no backup.')
                continue
            try:
                live_obj = manager.filter(pk=reg_pk).first()
                if live_obj:
                    aplicar_restore(live_obj, old_dict, todos_campos)
                    atualizados += 1
                else:
                    recriar_registro(model, old_dict)
                    criados += 1
                logger.info(
                    "[BACKUP RESTORE LOTE] %s restaurou %s #%s via backup #%s",
                    request.user.username, model.__name__, reg_pk, execucao.pk,
                )
            except Exception as exc:
                erros.append(f'ID {reg_pk}: {exc}')
    except Exception as exc:
        messages.error(request, f'Erro ao ler backup: {exc}')
    finally:
        if tempdb:
            dropar_temp(tempdb)
        if arquivo_temp and os.path.exists(arquivo_temp):
            try:
                os.remove(arquivo_temp)
            except Exception:
                pass

    partes = []
    if criados:
        partes.append(f'{criados} registro(s) recriado(s)')
    if atualizados:
        partes.append(f'{atualizados} registro(s) atualizado(s)')
    if partes:
        messages.success(request, 'Restauração em lote concluída: ' + ', '.join(partes) + '.')
    if erros:
        messages.warning(request, 'Erros: ' + ' | '.join(erros))

    return redirect(redirect_url)


# ==========================================
# GERENCIADOR DE ARQUIVOS E TERMINAL DO SERVIDOR DE BACKUP
# ==========================================

def _resolver_arquivo_db(execucao) -> tuple[str | None, str | None]:
    """
    Retorna (caminho_para_usar, caminho_temp_ou_None).
    1. Se o arquivo local existe → (arquivo_local, None)
    2. Se não existe mas foi enviado ao remoto → baixa via SFTP para /tmp,
       retorna (caminho_tmp, caminho_tmp) — o chamador deve apagar o tmp ao final.
    3. Se nenhum dos dois → (None, None)
    """
    import tempfile

    arquivo_local = execucao.arquivo_db or ''
    if arquivo_local and os.path.exists(arquivo_local):
        return arquivo_local, None

    if not execucao.enviado_remoto or not arquivo_local:
        return None, None

    try:
        from . import backup_server as _bs
        destino = BackupDestino.get_instance()
        if not destino.host or not destino.usuario:
            return None, None

        nome_arquivo = os.path.basename(arquivo_local)
        caminho_remoto = f"{destino.diretorio_destino.rstrip('/')}/{nome_arquivo}"

        conteudo = _bs.baixar_arquivo(destino, caminho_remoto)
        sufixo = os.path.splitext(nome_arquivo)[-1]
        fd, tmp_path = tempfile.mkstemp(suffix=sufixo)
        with os.fdopen(fd, 'wb') as f:
            f.write(conteudo)
        return tmp_path, tmp_path
    except Exception as exc:
        logger.warning("_resolver_arquivo_db: falha ao baixar do remoto: %s", exc)
        return None, None


def _caminho_seguro(caminho: str) -> str:
    """Normaliza e impede directory traversal (../) no caminho informado pelo usuário."""
    import posixpath
    caminho = caminho or '/'
    if not caminho.startswith('/'):
        caminho = '/' + caminho
    return posixpath.normpath(caminho)


@login_required
def backup_arquivos(request):
    """Gerenciador de arquivos do servidor de backup via SFTP: navegar, ver, editar, excluir."""
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from . import backup_server

    destino = BackupDestino.get_instance()
    caminho = _caminho_seguro(request.GET.get('path', destino.diretorio_destino or '/'))
    erro = None
    itens = None

    pasta_backup = (destino.diretorio_destino or '').rstrip('/')
    eh_pasta_backup = (caminho.rstrip('/') == pasta_backup)
    meses = None
    sem_data = None

    if not destino.host or not destino.usuario:
        erro = 'Servidor de backup não configurado (host/usuário ausentes).'
    else:
        try:
            itens = backup_server.listar_diretorio(destino, caminho)
            if eh_pasta_backup:
                meses, sem_data = backup_server.agrupar_por_data(itens)
        except Exception as exc:
            erro = f'Erro ao listar diretório: {exc}'

    pai = caminho.rsplit('/', 1)[0] or '/'

    return render(request, 'informatica/backup_arquivos.html', {
        'destino': destino,
        'caminho': caminho,
        'pai': pai,
        'itens': itens,
        'erro': erro,
        'meses': meses,
        'sem_data': sem_data,
        'eh_pasta_backup': eh_pasta_backup,
    })


@login_required
def backup_arquivo_ver(request):
    """Visualiza/edita um arquivo de texto do servidor de backup."""
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from . import backup_server

    destino = BackupDestino.get_instance()
    caminho = _caminho_seguro(request.GET.get('path') or request.POST.get('path', ''))
    erro = None
    conteudo = None
    pasta = caminho.rsplit('/', 1)[0] or '/'
    eh_texto = backup_server.eh_extensao_texto(caminho)

    if request.method == 'POST':
        novo_conteudo = request.POST.get('conteudo', '')
        try:
            backup_server.salvar_arquivo_texto(destino, caminho, novo_conteudo)
            registrar(
                request.user, secao='informatica',
                permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
                acao='editou', descricao=f"editou arquivo {caminho} no servidor de backup",
                objeto_tipo='ArquivoBackup', objeto_id=caminho,
            )
            messages.success(request, 'Arquivo salvo com sucesso.')
            return redirect(f"{reverse('informatica:backup_arquivo_ver')}?path={caminho}")
        except Exception as exc:
            erro = f'Erro ao salvar: {exc}'
    elif eh_texto:
        try:
            conteudo = backup_server.ler_arquivo_texto(destino, caminho)
        except Exception as exc:
            erro = f'Erro ao ler arquivo: {exc}'

    return render(request, 'informatica/backup_arquivo_ver.html', {
        'destino': destino,
        'caminho': caminho,
        'pasta': pasta,
        'conteudo': conteudo,
        'eh_texto': eh_texto,
        'erro': erro,
    })


@login_required
def backup_arquivo_baixar(request):
    """Baixa um arquivo do servidor de backup."""
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from django.http import HttpResponse, Http404
    from . import backup_server

    destino = BackupDestino.get_instance()
    caminho = _caminho_seguro(request.GET.get('path', ''))
    if not caminho or caminho == '/':
        raise Http404()

    try:
        conteudo = backup_server.baixar_arquivo(destino, caminho)
    except Exception as exc:
        messages.error(request, f'Erro ao baixar arquivo: {exc}')
        return redirect(f"{reverse('informatica:backup_arquivos')}?path={caminho.rsplit('/', 1)[0]}")

    nome_arquivo = caminho.rsplit('/', 1)[-1]
    resp = HttpResponse(conteudo, content_type='application/octet-stream')
    resp['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return resp


@login_required
@require_POST
def backup_arquivo_excluir(request):
    """Exclui um arquivo (ou pasta vazia) do servidor de backup."""
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from . import backup_server

    destino = BackupDestino.get_instance()
    caminho = _caminho_seguro(request.POST.get('path', ''))
    is_dir = request.POST.get('is_dir') == '1'
    pasta = caminho.rsplit('/', 1)[0] or '/'

    try:
        if is_dir:
            backup_server.excluir_diretorio_vazio(destino, caminho)
        else:
            backup_server.excluir_arquivo(destino, caminho)
        registrar(
            request.user, secao='informatica',
            permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
            acao='excluiu', descricao=f"excluiu {'pasta' if is_dir else 'arquivo'} {caminho} no servidor de backup",
            objeto_tipo='ArquivoBackup', objeto_id=caminho,
        )
        messages.success(request, f'"{caminho.rsplit("/", 1)[-1]}" excluído com sucesso.')
    except Exception as exc:
        messages.error(request, f'Erro ao excluir: {exc}')

    return redirect(f"{reverse('informatica:backup_arquivos')}?path={pasta}")


@login_required
def backup_terminal(request):
    """Página do terminal SSH interativo do servidor de backup (via WebSocket)."""
    if not is_informatica_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    destino = BackupDestino.get_instance()
    return render(request, 'informatica/backup_terminal.html', {'destino': destino})


# ==========================================
# MÓDULO GESTÃO DE MATERIAIS E CAUTELAS
# ==========================================
@login_required
def gestao_materiais_view(request):
    if not is_informatica_secao(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    militares = Efetivo.objects.all().order_by('posto', 'nome_guerra')
    
    militares_info = Efetivo.objects.filter(
        Q(setor__icontains='informática') | Q(subsetor__icontains='informática') |
        Q(setor__icontains='informatica') | Q(subsetor__icontains='informatica')
    ).order_by('posto', 'nome_guerra')

    grupos = GrupoMaterial.objects.all().order_by('nome')
    subgrupos = SubgrupoMaterial.objects.all().select_related('grupo').order_by('grupo__nome', 'nome')
    
    # 1. Busque os setores
    setores = Setor.objects.all().order_by('nome')

    # 2. Atualize a query de materiais para incluir 'secao' no select_related
    materiais = Material.objects.all().select_related('subgrupo__grupo', 'prateleira__armario', 'secao').order_by('nome')
    
    # Busca armários para o painel de armários e selects
    armarios = Armario.objects.all().prefetch_related('prateleiras__materiais').order_by('nome')
    
    cautelas_ativas = Cautela.objects.filter(ativa=True).select_related('sobreaviso', 'recebedor').order_by('-data_emissao')
    _historico_qs = Cautela.objects.filter(ativa=False).select_related('sobreaviso', 'recebedor').order_by('-data_emissao')
    _historico_total = _historico_qs.count()
    cautelas_historico = _historico_qs[:100]
    cautelas_historico_has_more = _historico_total > 100

    materiais_disponiveis = materiais.filter(quantidade_disponivel__gt=0, funcionando=True)
    
    materiais_json = [{
        'id': mat.id, 'grupo_id': mat.subgrupo.grupo.id, 'grupo_nome': mat.subgrupo.grupo.nome,
        'nome': mat.nome, 'serial': mat.serial or '', 'quantidade_disponivel': mat.quantidade_disponivel,
        'atributos': mat.atributos_extras or {}
    } for mat in materiais_disponiveis]
    
    acervo_json = [{
        'id': mat.id, 
        'subgrupo_id': mat.subgrupo.id, 
        'subgrupo_nome': mat.subgrupo.nome,
        'grupo_id': mat.subgrupo.grupo.id,
        'grupo_nome': mat.subgrupo.grupo.nome,
        'nome': mat.nome,
        'codigo': mat.codigo or '', 'serial': mat.serial or '',
        'quantidade': mat.quantidade, 'quantidade_disponivel': mat.quantidade_disponivel,
        'funcionando': 1 if mat.funcionando else 0,
        'motivo_defeito': mat.motivo_defeito or '',
        'atributos': mat.atributos_extras or {},
        'prateleira_id': mat.prateleira.id if mat.prateleira else None,
        'prateleira_nome': mat.prateleira.nome if mat.prateleira else None,
        'armario_id': mat.prateleira.armario.id if mat.prateleira else None,
        'armario_nome': mat.prateleira.armario.nome if mat.prateleira else None,
        'secao_id': mat.secao.id if mat.secao else '',
        'secao_nome': mat.secao.nome if mat.secao else 'Geral/Estoque',
        'localizacao_texto': mat.localizacao_texto or '',
    } for mat in materiais]

    # Busca o telefone da última cautela de cada militar em 1 query via Subquery (evita N+1)
    from django.db.models import OuterRef, Subquery as _Subquery
    _last_tel = Cautela.objects.filter(
        recebedor=OuterRef('pk')
    ).order_by('-data_emissao').values('telefone_contato')[:1]
    militares_com_tel = militares.annotate(_last_telefone=_Subquery(_last_tel))
    militares_dados_json = [
        {'id': m.id, 'telefone': m._last_telefone or '', 'saram': getattr(m, 'saram', '')}
        for m in militares_com_tel
    ]

    militares_info_json = []
    for m in militares_info:
        militares_info_json.append({
            'id': m.id,
            'assinatura': getattr(m, 'assinatura', None)
        })
        
    armarios_json = [{
        'id': arm.id, 'nome': arm.nome, 'localizacao': arm.localizacao or '',
        'prateleiras': [{'id': p.id, 'nome': p.nome} for p in arm.prateleiras.all()]
    } for arm in armarios]

    # 4. Adicione os setores ao contexto
    context = {
        'page_title': 'Gestão de Materiais e Cautelas',
        'militares': militares,
        'militares_info': militares_info,
        'grupos': grupos, 'subgrupos': subgrupos, 'materiais': materiais,
        'armarios': armarios,
        'setores': setores,
        'cautelas_ativas': cautelas_ativas,
        'cautelas_historico': cautelas_historico,
        'cautelas_historico_has_more': cautelas_historico_has_more,
        'materiais_json': json.dumps(materiais_json),
        'acervo_json': json.dumps(acervo_json),
        'militares_dados_json': json.dumps(militares_dados_json),
        'militares_info_json': json.dumps(militares_info_json),
        'armarios_json': json.dumps(armarios_json),
    }
    return render(request, 'informatica/gestao_materiais.html', context)


@staff_member_required
@require_POST
def api_add_grupo(request):
    data = json.loads(request.body)
    try:
        GrupoMaterial.objects.create(nome=data.get('nome'))
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_add_subgrupo(request):
    data = json.loads(request.body)
    try:
        grupo = GrupoMaterial.objects.get(id=data.get('grupo_id'))
        SubgrupoMaterial.objects.create(grupo=grupo, nome=data.get('nome'))
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_add_armario(request):
    data = json.loads(request.body)
    try:
        Armario.objects.create(nome=data.get('nome'), localizacao=data.get('localizacao', ''))
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_add_prateleira(request):
    data = json.loads(request.body)
    try:
        armario = Armario.objects.get(id=data.get('armario_id'))
        Prateleira.objects.create(armario=armario, nome=data.get('nome'))
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_edit_armario(request, pk):
    data = json.loads(request.body)
    try:
        armario = Armario.objects.get(pk=pk)
        armario.nome = data.get('nome')
        armario.localizacao = data.get('localizacao', '')
        armario.save()
        return JsonResponse({'status': 'success'})
    except Exception as e: 
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_delete_armario(request, pk):
    try:
        armario = Armario.objects.get(pk=pk)
        armario.delete() # Excluirá prateleiras (CASCADE). Materiais ficarão com local = NULL (SET_NULL)
        return JsonResponse({'status': 'success'})
    except Exception as e: 
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_edit_prateleira(request, pk):
    data = json.loads(request.body)
    try:
        prateleira = Prateleira.objects.get(pk=pk)
        prateleira.nome = data.get('nome')
        prateleira.save()
        return JsonResponse({'status': 'success'})
    except Exception as e: 
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_delete_prateleira(request, pk):
    try:
        prateleira = Prateleira.objects.get(pk=pk)
        prateleira.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e: 
        return JsonResponse({'status': 'error', 'message': str(e)})


@staff_member_required
@require_POST
def api_add_material(request):
    data = json.loads(request.body)
    try:
        subgrupo = SubgrupoMaterial.objects.get(id=data.get('subgrupo_id'))
        prateleira_id = data.get('prateleira_id')
        prateleira = Prateleira.objects.get(id=prateleira_id) if prateleira_id else None
        
        secao_id = data.get('secao_id')
        secao = Setor.objects.get(id=secao_id) if secao_id else None
        
        serial = data.get('serial')
        qtd = int(data.get('quantidade', 1))
        
        if serial and Material.objects.filter(subgrupo__grupo=subgrupo.grupo, serial=serial).exists():
            return JsonResponse({'status': 'error', 'message': f"O serial '{serial}' já está em uso nesta categoria."})

        Material.objects.create(
            subgrupo=subgrupo,
            secao=secao,
            nome=data.get('nome'),
            codigo=data.get('codigo'),
            serial=serial,
            prateleira=prateleira,
            localizacao_texto=data.get('localizacao_texto', ''),
            quantidade=qtd,
            quantidade_disponivel=qtd,
            funcionando=data.get('funcionando', True),
            motivo_defeito=data.get('motivo_defeito', ''),
            atributos_extras=data.get('atributos_extras', {})
        )
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_edit_material(request, pk):
    data = json.loads(request.body)
    try:
        material = Material.objects.get(pk=pk)
        subgrupo = SubgrupoMaterial.objects.get(id=data.get('subgrupo_id'))
        prateleira_id = data.get('prateleira_id')
        prateleira = Prateleira.objects.get(id=prateleira_id) if prateleira_id else None
        
        secao_id = data.get('secao_id')
        secao = Setor.objects.get(id=secao_id) if secao_id else None
        
        serial = data.get('serial')
        qtd = int(data.get('quantidade', 1))
        
        if serial and serial != material.serial and Material.objects.filter(subgrupo__grupo=subgrupo.grupo, serial=serial).exists():
            return JsonResponse({'status': 'error', 'message': f"O serial '{serial}' já está em uso nesta categoria."})

        diff = qtd - material.quantidade
        nova_disp = material.quantidade_disponivel + diff
        
        if nova_disp < 0:
            return JsonResponse({'status': 'error', 'message': 'A quantidade total não pode ser menor que a quantidade que já está emprestada!'})
            
        material.subgrupo = subgrupo
        material.secao = secao
        material.nome = data.get('nome')
        material.codigo = data.get('codigo')
        material.serial = serial
        material.prateleira = prateleira
        material.localizacao_texto = data.get('localizacao_texto', '')
        material.quantidade = qtd
        material.quantidade_disponivel = nova_disp
        material.funcionando = data.get('funcionando', True)
        material.motivo_defeito = data.get('motivo_defeito', '')
        material.atributos_extras = data.get('atributos_extras', {})
        
        if material.quantidade_disponivel == 0:
            material.disponivel = False
        else:
            material.disponivel = True
            
        material.save()
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_delete_material(request, pk):
    data = json.loads(request.body)
    password = data.get('password')
    
    # Autenticação dupla para excluir materiais
    user = authenticate(username=request.user.username, password=password)
    if user is None:
        return JsonResponse({'status': 'error', 'message': 'Senha incorreta. A exclusão não foi autorizada.'})
        
    try:
        material = Material.objects.get(pk=pk)
        material.delete()
        return JsonResponse({'status': 'success'})
    except ProtectedError:
        return JsonResponse({'status': 'error', 'message': 'Este material não pode ser excluído pois faz parte do histórico de uma ou mais cautelas. Sugestão: Marque-o como "Com Defeito" ou Inoperante.'})
    except Exception as e: 
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_delete_grupo(request, pk):
    try:
        grupo = GrupoMaterial.objects.get(pk=pk)
        if Material.objects.filter(subgrupo__grupo=grupo).exists():
            return JsonResponse({'status': 'error', 'message': 'Existem materiais cadastrados neste grupo.'})
        grupo.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_delete_subgrupo(request, pk):
    try:
        subgrupo = SubgrupoMaterial.objects.get(pk=pk)
        if subgrupo.materiais.exists():
            return JsonResponse({'status': 'error', 'message': 'Existem materiais cadastrados neste subgrupo.'})
        subgrupo.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})


@staff_member_required
@require_POST
def api_edit_grupo(request, pk):
    data = json.loads(request.body)
    try:
        grupo = GrupoMaterial.objects.get(pk=pk)
        nome = data.get('nome', '').strip()
        if not nome:
            return JsonResponse({'status': 'error', 'message': 'Nome não pode ser vazio.'})
        grupo.nome = nome
        grupo.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_edit_subgrupo(request, pk):
    data = json.loads(request.body)
    try:
        subgrupo = SubgrupoMaterial.objects.get(pk=pk)
        nome = data.get('nome', '').strip()
        if not nome:
            return JsonResponse({'status': 'error', 'message': 'Nome não pode ser vazio.'})
        subgrupo.nome = nome
        subgrupo.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_salvar_cautela(request):
    data = json.loads(request.body)
    try:
        sobreaviso = Efetivo.objects.get(id=data.get('sobreaviso_id'))
        recebedor = Efetivo.objects.get(id=data.get('recebedor_id'))
        materiais_list = data.get('materiais', [])

        if not materiais_list:
            return JsonResponse({'status': 'error', 'message': 'Nenhum material selecionado.'})

        with transaction.atomic():
            if data.get('salvar_padrao') and data.get('assinatura_sobreaviso'):
                sobreaviso.assinatura = data.get('assinatura_sobreaviso')
                sobreaviso.save()

            cautela = Cautela.objects.create(
                sobreaviso=sobreaviso,
                recebedor=recebedor,
                assinatura_sobreaviso=data.get('assinatura_sobreaviso'),
                assinatura_recebedor=data.get('assinatura_recebedor'),
                nome_missao=data.get('nome_missao', ''),
                telefone_contato=data.get('telefone_contato', '')
            )

            mat_ids = [int(m['id']) for m in materiais_list]
            # select_for_update impede que dois requests simultâneos emprestem o mesmo estoque
            materiais_map = {
                m.pk: m
                for m in Material.objects.select_for_update().filter(pk__in=mat_ids)
            }

            for mat_data in materiais_list:
                material = materiais_map[int(mat_data['id'])]
                qtd = int(mat_data['qtd'])
                if material.quantidade_disponivel < qtd:
                    raise ValueError(f"Material {material.nome} só tem {material.quantidade_disponivel} disponíveis!")
                CautelaItem.objects.create(cautela=cautela, material=material, quantidade=qtd)
                material.quantidade_disponivel -= qtd
                if material.quantidade_disponivel == 0:
                    material.disponivel = False
                material.save()

        return JsonResponse({'status': 'success', 'cautela_id': cautela.id})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_devolver_cautela(request, pk):
    data = json.loads(request.body)
    try:
        with transaction.atomic():
            cautela = Cautela.objects.select_for_update().get(id=pk)
            sobreaviso_devolucao = Efetivo.objects.get(id=data.get('sobreaviso_id'))
            agora = timezone.now()

            cautela.ativa = False
            cautela.data_devolucao = agora
            cautela.recebedor_devolucao = sobreaviso_devolucao
            cautela.assinatura_devolucao = data.get('assinatura_devolucao')
            cautela.save()

            for item in cautela.itens.select_related('material').filter(devolvido=False):
                item.devolvido = True
                item.data_devolucao = agora
                item.recebedor_devolucao = sobreaviso_devolucao
                item.assinatura_devolucao = data.get('assinatura_devolucao')
                item.save()
                item.material.quantidade_disponivel += item.quantidade
                item.material.disponivel = True
                item.material.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_devolver_item_cautela(request, item_id):
    data = json.loads(request.body)
    try:
        with transaction.atomic():
            item = CautelaItem.objects.select_related('material', 'cautela').select_for_update().get(id=item_id)
            sobreaviso_devolucao = Efetivo.objects.get(id=data.get('sobreaviso_id'))
            agora = timezone.now()

            item.devolvido = True
            item.data_devolucao = agora
            item.recebedor_devolucao = sobreaviso_devolucao
            item.assinatura_devolucao = data.get('assinatura_devolucao')
            item.save()

            item.material.quantidade_disponivel += item.quantidade
            item.material.disponivel = True
            item.material.save()

            cautela = item.cautela
            if not cautela.itens.filter(devolvido=False).exists():
                cautela.ativa = False
                cautela.data_devolucao = agora
                cautela.recebedor_devolucao = sobreaviso_devolucao
                cautela.assinatura_devolucao = data.get('assinatura_devolucao')
                cautela.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@staff_member_required
@require_POST
def api_devolver_multiplos_itens(request, cautela_id):
    data = json.loads(request.body)
    try:
        with transaction.atomic():
            cautela = Cautela.objects.select_for_update().get(id=cautela_id)
            sobreaviso_devolucao = Efetivo.objects.get(id=data.get('sobreaviso_id'))
            item_ids = data.get('item_ids', [])
            agora = timezone.now()

            itens = (
                CautelaItem.objects.select_related('material')
                .select_for_update()
                .filter(id__in=item_ids, cautela=cautela, devolvido=False)
            )
            for item in itens:
                item.devolvido = True
                item.data_devolucao = agora
                item.recebedor_devolucao = sobreaviso_devolucao
                item.assinatura_devolucao = data.get('assinatura_devolucao')
                item.save()

                item.material.quantidade_disponivel += item.quantidade
                item.material.disponivel = True
                item.material.save()

            if not cautela.itens.filter(devolvido=False).exists():
                cautela.ativa = False
                cautela.data_devolucao = agora
                cautela.recebedor_devolucao = sobreaviso_devolucao
                cautela.assinatura_devolucao = data.get('assinatura_devolucao')
                cautela.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@staff_member_required
@require_POST
def api_add_item_cautela(request, cautela_id):
    data = json.loads(request.body)
    try:
        cautela = Cautela.objects.get(id=cautela_id)
        material = Material.objects.get(id=data.get('material_id'))
        qtd = int(data.get('quantidade', 1))
        
        if material.quantidade_disponivel < qtd:
            return JsonResponse({'status': 'error', 'message': f'Estoque insuficiente. Apenas {material.quantidade_disponivel} disponíveis.'})
            
        CautelaItem.objects.create(cautela=cautela, material=material, quantidade=qtd)
        
        material.quantidade_disponivel -= qtd
        if material.quantidade_disponivel == 0:
            material.disponivel = False
        material.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

from django.http import HttpResponse

@staff_member_required
def exportar_armarios_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Erro: A biblioteca 'openpyxl' não está instalada. Execute 'pip install openpyxl' no ambiente do servidor.", status=500)

    armarios_ids = request.GET.get('armarios', '')
    is_completo = request.GET.get('completo') == 'true'

    query = Armario.objects.all().prefetch_related('prateleiras__materiais')
    
    if armarios_ids:
        ids = [int(i) for i in armarios_ids.split(',') if i.isdigit()]
        if ids:
            query = query.filter(id__in=ids)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventário de Armários"
    
    # Estilos de Formatação
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(bold=True, size=16, color="1F4E78")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Título do Relatório (Linha 1)
    max_col = 10 if is_completo else 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1, value=f"Relatório de Inventário de Armários - {'Completo' if is_completo else 'Simplificado'} ({timezone.now().strftime('%d/%m/%Y')})")
    title_cell.font = title_font
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 30
    
    # Cabeçalhos (Linha 2)
    if is_completo:
        headers = ['Armário', 'Localização', 'Prateleira', 'Material', 'S/N', 'Código Interno', 'Qtd', 'Status', 'Atributos Técnicos', 'Observações/Defeito']
    else:
        headers = ['Armário', 'Localização', 'Prateleira', 'Material', 'S/N', 'Código Interno', 'Qtd', 'Status']
        
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    current_row = 3
    for armario in query:
        prateleiras = armario.prateleiras.all()
        if not prateleiras:
            row_data = [armario.nome, armario.localizacao or '-', 'Sem Prateleiras', '-', '-', '-', '-', '-']
            if is_completo: row_data.extend(['-', '-'])
            ws.append(row_data)
            for col_num in range(1, len(row_data) + 1):
                ws.cell(row=current_row, column=col_num).border = thin_border
                ws.cell(row=current_row, column=col_num).alignment = align_center
            current_row += 1
            continue
            
        for prateleira in prateleiras:
            materiais = prateleira.materiais.all()
            if not materiais:
                row_data = [armario.nome, armario.localizacao or '-', prateleira.nome, 'Vazia', '-', '-', '-', '-']
                if is_completo: row_data.extend(['-', '-'])
                ws.append(row_data)
                for col_num in range(1, len(row_data) + 1):
                    ws.cell(row=current_row, column=col_num).border = thin_border
                    ws.cell(row=current_row, column=col_num).alignment = align_center
                current_row += 1
                continue
                
            for mat in materiais:
                status = 'Funcionando' if mat.funcionando else 'Com Defeito'
                row_data = [
                    armario.nome,
                    armario.localizacao or '-',
                    prateleira.nome,
                    mat.nome,
                    mat.serial or '-',
                    mat.codigo or '-',
                    mat.quantidade,
                    status
                ]
                
                if is_completo:
                    attrs_str = "-"
                    if mat.atributos_extras:
                        # Junta os atributos com quebras de linha
                        attrs_str = "\n".join([f"• {k}: {v}" for k, v in mat.atributos_extras.items()])
                    row_data.append(attrs_str)
                    row_data.append(mat.motivo_defeito or '-')
                    
                ws.append(row_data)
                
                # Aplica estilos nas células inseridas
                for col_num, val in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.border = thin_border
                    # Alinha à esquerda apenas colunas de texto descritivo
                    if col_num in [4, 9, 10]:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                
                current_row += 1
                
    # Auto-ajuste de largura das colunas
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            try:
                # Ignora a linha do título principal na contagem de largura
                if cell.row > 1: 
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
            except:
                pass
        
        # Define largura máxima de 45 para não criar colunas absurdamente gigantes
        adjusted_width = min(max_length + 2, 45) 
        ws.column_dimensions[column].width = adjusted_width

    # Adiciona Filtros (Setinhas) na tabela (A2:J[ultima_linha])
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{current_row - 1}"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "Inventario_Completo.xlsx" if is_completo else "Inventario_Simplificado.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@staff_member_required
def imprimir_cautela(request, pk):
    cautela = get_object_or_404(Cautela, id=pk)
    return render(request, 'informatica/cautela_print.html', {'cautela': cautela})


@login_required
def configuracao_secoes(request):
    if not is_informatica_admin(request.user):
        return redirect('informatica:gestao_materiais')

    from Ouvidoria.models import PATD, Configuracao
    from Secao_pessoal.models import Efetivo
    from django.contrib.auth.models import User

    secoes = [
        {
            'nome': 'Ouvidoria',
            'icon': 'fa-file-alt',
            'cor': '#3b82f6',
            'stats': [
                {'label': 'PATDs Totais', 'value': PATD.objects.count()},
                {'label': 'Em Andamento', 'value': PATD.objects.exclude(status='finalizado').count()},
            ],
            'links': [
                {'label': 'Listar PATDs', 'url': 'Ouvidoria:index', 'icon': 'fa-list'},
                {'label': 'Configurações', 'url': 'informatica:configuracao_edit', 'icon': 'fa-cog'},
            ],
        },
        {
            'nome': 'Seção de Pessoal (S1)',
            'icon': 'fa-users',
            'cor': '#22c55e',
            'stats': [
                {'label': 'Militares', 'value': Efetivo.objects.count()},
                {'label': 'Utilizadores', 'value': User.objects.count()},
            ],
            'links': [
                {'label': 'Gestão de Efetivo', 'url': 'Secao_pessoal:index', 'icon': 'fa-users'},
                {'label': 'Chamada', 'url': 'chamada:chamada_index', 'icon': 'fa-check-square'},
            ],
        },
        {
            'nome': 'Seção de Operações',
            'icon': 'fa-cogs',
            'cor': '#f59e0b',
            'stats': [
                {'label': 'Missões', 'value': __import__('Secao_operacoes.models', fromlist=['Missao']).Missao.objects.count()},
            ],
            'links': [
                {'label': 'Acessar Seção', 'url': 'Secao_operacoes:index', 'icon': 'fa-arrow-right'},
                {'label': 'Missões (OMIS)', 'url': 'Secao_operacoes:missao_list', 'icon': 'fa-clipboard-list'},
                {'label': 'Configurações de Operações', 'url': 'Secao_operacoes:config_operacoes', 'icon': 'fa-cog'},
            ],
        },
        {
            'nome': 'Informática',
            'icon': 'fa-desktop',
            'cor': '#8b5cf6',
            'stats': [
                {'label': 'Materiais', 'value': Material.objects.count()},
                {'label': 'Em Cautela', 'value': Cautela.objects.filter(ativa=True).count()},
            ],
            'links': [
                {'label': 'Gestão de Materiais', 'url': 'informatica:gestao_materiais', 'icon': 'fa-boxes-stacked'},
                {'label': 'Utilizadores', 'url': 'informatica:user_list', 'icon': 'fa-users'},
                {'label': 'Grupos', 'url': 'informatica:group_list', 'icon': 'fa-user-tag'},
                {'label': 'Comandantes', 'url': 'informatica:configuracao_comandantes', 'icon': 'fa-star'},
            ],
        },
    ]

    return render(request, 'informatica/configuracao_secoes.html', {
        'secoes': secoes,
    })


# ==========================================
# PAINEL ADMIN OUVIDORIA
# ==========================================

@staff_member_required
def ouvidoria_admin_search(request):
    """Pesquisa PATDs pelo número ou nome do militar."""
    q = request.GET.get('q', '').strip()
    patds = PATD.all_objects.select_related('militar', 'oficial_responsavel')
    if q:
        patds = patds.filter(
            Q(numero_patd__icontains=q) |
            Q(militar__nome_guerra__icontains=q) |
            Q(militar__nome_completo__icontains=q)
        )
    patds = patds.order_by('-data_inicio')[:50]

    data = [{
        'id': p.id,
        'numero_patd': p.numero_patd,
        'militar': str(p.militar),
        'status': p.status,
        'status_display': p.get_status_display(),
        'deleted': p.deleted,
    } for p in patds]
    return JsonResponse(data, safe=False)


@staff_member_required
def ouvidoria_admin_patd_detail(request, pk):
    """Retorna detalhes completos de uma PATD para o painel admin."""
    patd = get_object_or_404(PATD.all_objects, pk=pk)

    # Calcula prazo para status relevantes de defesa
    prazo_info = None
    if patd.status in ('aguardando_justificativa', 'prazo_expirado') and patd.data_ciencia:
        from Ouvidoria.models import Configuracao as OuvidoriaConfig
        config = OuvidoriaConfig.load()
        from datetime import timedelta
        data_final = patd.data_ciencia
        dias_adicionados = 0
        while dias_adicionados < config.prazo_defesa_dias:
            data_final += timedelta(days=1)
            if data_final.weekday() < 5:
                dias_adicionados += 1
        calculated_deadline = (data_final + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        # Se houver override, usa ele como deadline efetivo
        deadline = patd.prazo_override if patd.prazo_override else calculated_deadline
        agora = timezone.now()
        restante = deadline - agora
        prazo_info = {
            'deadline': calculated_deadline.isoformat(),  # sempre o calculado (para exibição)
            'effective_deadline': deadline.isoformat(),    # efetivo (override ou calculado)
            'restante_segundos': int(restante.total_seconds()),
            'expirado': restante.total_seconds() < 0,
        }

    # Anexos
    anexos = [{
        'id': a.id,
        'tipo': a.tipo,
        'tipo_display': a.get_tipo_display(),
        'nome': os.path.basename(a.arquivo.name) if a.arquivo else '',
        'url': a.arquivo.url if a.arquivo else '',
        'data_upload': a.data_upload.isoformat(),
    } for a in patd.anexos.all().order_by('-data_upload')]

    # Documentos / assinaturas geradas
    docs_gerados = []
    if patd.assinatura_oficial:
        docs_gerados.append({'campo': 'assinatura_oficial', 'label': 'Assinatura do Oficial', 'url': patd.assinatura_oficial.url})
    if patd.assinatura_testemunha1:
        docs_gerados.append({'campo': 'assinatura_testemunha1', 'label': 'Assinatura Testemunha 1', 'url': patd.assinatura_testemunha1.url})
    if patd.assinatura_testemunha2:
        docs_gerados.append({'campo': 'assinatura_testemunha2', 'label': 'Assinatura Testemunha 2', 'url': patd.assinatura_testemunha2.url})
    if patd.assinatura_alegacao_defesa:
        docs_gerados.append({'campo': 'assinatura_alegacao_defesa', 'label': 'Assinatura Alegação de Defesa', 'url': patd.assinatura_alegacao_defesa.url})
    if patd.assinatura_reconsideracao:
        docs_gerados.append({'campo': 'assinatura_reconsideracao', 'label': 'Assinatura Reconsideração', 'url': patd.assinatura_reconsideracao.url})
    if patd.assinaturas_militar:
        sigs_validas = [s for s in (patd.assinaturas_militar or []) if s]
        if sigs_validas:
            docs_gerados.append({'campo': 'assinaturas_militar', 'label': f'Assinaturas do Militar ({len(sigs_validas)})', 'url': None})
    if patd.alegacao_defesa:
        docs_gerados.append({'campo': 'alegacao_defesa', 'label': 'Alegação de Defesa (texto)', 'url': None})
    if patd.texto_relatorio:
        docs_gerados.append({'campo': 'texto_relatorio', 'label': 'Relatório de Apuração', 'url': None})
    if patd.relatorio_final:
        docs_gerados.append({'campo': 'relatorio_final', 'label': 'Relatório Final', 'url': None})

    # Oficiais disponíveis (para troca)
    oficiais = list(
        Efetivo.objects.filter(oficial=True)
        .exclude(assinatura__isnull=True).exclude(assinatura__exact='')
        .order_by('posto', 'nome_guerra')
        .values('id', 'posto', 'nome_guerra')
    )

    status_choices = [{'value': k, 'label': v} for k, v in PATD.STATUS_CHOICES]

    data = {
        'id': patd.id,
        'numero_patd': patd.numero_patd,
        'militar': str(patd.militar),
        'oficial_responsavel_id': patd.oficial_responsavel_id,
        'oficial_responsavel': str(patd.oficial_responsavel) if patd.oficial_responsavel else None,
        'status': patd.status,
        'status_display': patd.get_status_display(),
        'data_ciencia': patd.data_ciencia.isoformat() if patd.data_ciencia else None,
        'prazo_override': patd.prazo_override.isoformat() if patd.prazo_override else None,
        'data_inicio': patd.data_inicio.isoformat() if patd.data_inicio else None,
        'justificado': patd.justificado,
        'deleted': patd.deleted,
        'arquivado': patd.arquivado,
        'prazo_info': prazo_info,
        'anexos': anexos,
        'docs_gerados': docs_gerados,
        'oficiais_disponiveis': oficiais,
        'status_choices': status_choices,
    }
    return JsonResponse(data)


@staff_member_required
@require_POST
def ouvidoria_admin_update(request, pk):
    """Aplica alterações administrativas a uma PATD."""
    patd = get_object_or_404(PATD.all_objects, pk=pk)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)

    action = data.get('action')
    changes = []

    if action == 'change_status':
        novo_status = data.get('status')
        valid_statuses = [k for k, _ in PATD.STATUS_CHOICES]
        if novo_status not in valid_statuses:
            return JsonResponse({'status': 'error', 'message': 'Status inválido.'}, status=400)
        patd.status = novo_status
        patd.save(update_fields=['status'])
        changes.append(f'Status → {patd.get_status_display()}')

    elif action == 'set_prazo_override':
        # Define o deadline de defesa diretamente, sem tocar em data_ciencia
        prazo_str = data.get('prazo_override')
        if prazo_str == '' or prazo_str is None:
            # Limpar override (volta ao cálculo normal)
            patd.prazo_override = None
            patd.save(update_fields=['prazo_override'])
            changes.append('Prazo override → removido (volta ao cálculo normal)')
        else:
            try:
                from datetime import datetime as dt
                nova_data = timezone.make_aware(dt.fromisoformat(prazo_str))
                patd.prazo_override = nova_data
                patd.save(update_fields=['prazo_override'])
                changes.append(f'Prazo de defesa → {nova_data.strftime("%d/%m/%Y %H:%M")}')
            except (ValueError, TypeError):
                return JsonResponse({'status': 'error', 'message': 'Formato de data inválido.'}, status=400)

    elif action == 'change_oficial':
        oficial_id = data.get('oficial_id')
        if oficial_id:
            oficial = get_object_or_404(Efetivo, pk=oficial_id, oficial=True)
            patd.oficial_responsavel = oficial
            # Não usar update_fields aqui para que o save() do modelo
            # possa atualizar o status para 'aguardando_aprovacao_atribuicao' automaticamente
            patd.save()
            changes.append(f'Oficial → {oficial} (aguardando aceitação)')
        else:
            patd.oficial_responsavel = None
            patd.save()
            changes.append('Oficial → Removido')

    elif action == 'delete_field':
        # Apaga um campo de assinatura/documento
        campo = data.get('campo')
        campos_permitidos = [
            'assinatura_oficial', 'assinatura_testemunha1', 'assinatura_testemunha2',
            'assinatura_alegacao_defesa', 'assinatura_reconsideracao',
            'alegacao_defesa', 'texto_relatorio', 'relatorio_final', 'assinaturas_militar',
        ]
        if campo not in campos_permitidos:
            return JsonResponse({'status': 'error', 'message': 'Campo não permitido.'}, status=400)
        field = getattr(patd, campo)
        if hasattr(field, 'delete'):  # FileField
            field.delete(save=False)
        setattr(patd, campo, None if campo not in ('alegacao_defesa', 'texto_relatorio', 'relatorio_final') else '')
        if campo == 'assinaturas_militar':
            patd.assinaturas_militar = []
        patd.save(update_fields=[campo])
        changes.append(f'Campo {campo} → removido')

    else:
        return JsonResponse({'status': 'error', 'message': 'Ação desconhecida.'}, status=400)

    logger.info(f"[ADMIN OUVIDORIA] PATD {patd.numero_patd} — {'; '.join(changes)} por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='editou', descricao=f"editou a PATD {patd.numero_patd} via painel admin ({'; '.join(changes)})",
        objeto_tipo='PATD', objeto_id=patd.numero_patd,
    )
    return JsonResponse({'status': 'success', 'changes': changes})


@staff_member_required
@require_POST
def ouvidoria_admin_delete_anexo(request, patd_pk, anexo_pk):
    """Remove um anexo específico de uma PATD."""
    patd = get_object_or_404(PATD.all_objects, pk=patd_pk)
    anexo = get_object_or_404(Anexo, pk=anexo_pk, patd=patd)
    nome = os.path.basename(anexo.arquivo.name) if anexo.arquivo else str(anexo_pk)
    try:
        anexo.arquivo.delete(save=False)
    except Exception:
        pass
    anexo.delete()
    logger.info(f"[ADMIN OUVIDORIA] Anexo '{nome}' removido da PATD {patd.numero_patd} por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='excluiu', descricao=f"removeu o anexo '{nome}' da PATD {patd.numero_patd}",
        objeto_tipo='PATD', objeto_id=patd.numero_patd,
    )
    return JsonResponse({'status': 'success', 'message': f'Anexo "{nome}" removido.'})


# ==========================================
# LIXEIRA ADMIN
# ==========================================

@staff_member_required
@require_POST
def ouvidoria_lixeira_config(request):
    """Atualiza o tempo de retenção da lixeira."""
    try:
        dias = int(request.POST.get('dias_retencao_lixeira', 30))
        if dias < 1 or dias > 365:
            raise ValueError
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Valor inválido (1–365 dias).'}, status=400)
    config = Configuracao.load()
    config.dias_retencao_lixeira = dias
    config.save(update_fields=['dias_retencao_lixeira'])
    logger.info(f"[LIXEIRA] Retenção alterada para {dias} dias por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='configurou', descricao=f"alterou a retenção da lixeira de PATD para {dias} dias",
        objeto_tipo='Configuração',
    )
    return JsonResponse({'status': 'success', 'dias': dias})


@staff_member_required
@require_POST
def ouvidoria_lixeira_restore(request, pk):
    """Restaura uma PATD da lixeira."""
    patd = get_object_or_404(PATD.all_objects, pk=pk, deleted=True)
    numero = patd.numero_patd
    patd.deleted = False
    patd.deleted_at = None
    patd.save(update_fields=['deleted', 'deleted_at'])
    logger.info(f"[LIXEIRA] PATD {numero} restaurada por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='restaurou', descricao=f"restaurou a PATD {numero} da lixeira",
        objeto_tipo='PATD', objeto_id=numero,
    )
    return JsonResponse({'status': 'success', 'message': f'PATD {numero} restaurada.'})


@staff_member_required
@require_POST
def ouvidoria_lixeira_delete(request, pk):
    """Exclui permanentemente uma PATD da lixeira."""
    patd = get_object_or_404(PATD.all_objects, pk=pk, deleted=True)
    numero = patd.numero_patd
    patd.delete()
    logger.info(f"[LIXEIRA] PATD {numero} excluída permanentemente por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='excluiu', descricao=f"excluiu permanentemente a PATD {numero} da lixeira",
        objeto_tipo='PATD', objeto_id=numero,
    )
    return JsonResponse({'status': 'success', 'message': f'PATD {numero} excluída permanentemente.'})


@staff_member_required
@require_POST
def ouvidoria_lixeira_set_deleted_at(request, pk):
    """Altera a data/hora de entrada na lixeira de uma PATD específica."""
    patd = get_object_or_404(PATD.all_objects, pk=pk, deleted=True)
    try:
        body = json.loads(request.body)
        from datetime import datetime as dt
        nova_data = timezone.make_aware(dt.fromisoformat(body['deleted_at']))
    except (json.JSONDecodeError, KeyError, ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Formato de data inválido.'}, status=400)
    patd.deleted_at = nova_data
    patd.save(update_fields=['deleted_at'])
    logger.info(f"[LIXEIRA] deleted_at de PATD {patd.numero_patd} alterado para {nova_data} por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='editou', descricao=f"alterou a data de entrada na lixeira da PATD {patd.numero_patd} para {nova_data.strftime('%d/%m/%Y %H:%M')}",
        objeto_tipo='PATD', objeto_id=patd.numero_patd,
    )
    return JsonResponse({'status': 'success', 'message': f'Data atualizada para {nova_data.strftime("%d/%m/%Y %H:%M")}.'})


@staff_member_required
@require_POST
def ouvidoria_lixeira_esvaziar(request):
    """Exclui permanentemente todas as PATDs na lixeira."""
    patds = PATD.all_objects.filter(deleted=True)
    count = patds.count()
    patds.delete()
    logger.info(f"[LIXEIRA] {count} PATD(s) excluídas por {request.user.username}")
    registrar(
        request.user, secao='ouvidoria', permissao=resolver_label(request.user, _INFORMATICA_PERMISSAO_MAP),
        acao='excluiu', descricao=f"esvaziou a lixeira de PATD ({count} excluída(s) permanentemente)",
        objeto_tipo='PATD',
    )
    return JsonResponse({'status': 'success', 'count': count, 'message': f'{count} PATD(s) excluídas permanentemente.'})


# ==========================================
# LOGS DE ALTERAÇÃO DA ALEGAÇÃO DE DEFESA
# ==========================================

@login_required
def logs_alegacao_defesa(request):
    """Lista todos os logs de alteração da alegação de defesa para o painel informatica."""
    if not is_informatica_admin(request.user):
        return redirect('home:index')

    qs = AlegacaoDefesaLog.objects.select_related(
        'patd', 'patd__militar', 'usuario'
    ).order_by('-data_alteracao')

    patd_pk = request.GET.get('patd')
    if patd_pk:
        qs = qs.filter(patd__pk=patd_pk)

    return render(request, 'informatica/logs_alegacao_defesa.html', {
        'logs': qs[:200],
        'patd_pk_filtro': patd_pk or '',
    })


# ==========================================
# AUDITORIA (Fase 3) — página dedicada + APIs JSON
# ==========================================
@login_required
def auditoria_page(request):
    if not is_informatica_secao(request.user):
        return redirect('home:index')
    return render(request, 'informatica/auditoria.html')


@login_required
def auditoria_search(request):
    if not is_informatica_secao(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from auditoria.models import LogAuditoria

    qs = LogAuditoria.objects.select_related('usuario').all()
    usuario = request.GET.get('usuario', '').strip()
    secao = request.GET.get('secao', '').strip()
    acao = request.GET.get('acao', '').strip()
    objeto_tipo = request.GET.get('objeto_tipo', '').strip()
    permissao = request.GET.get('permissao', '').strip()
    busca = request.GET.get('busca', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()

    if usuario:
        qs = qs.filter(Q(username__icontains=usuario) | Q(nome_guerra__icontains=usuario))
    if secao:
        qs = qs.filter(secao=secao)
    if acao:
        qs = qs.filter(acao=acao)
    if objeto_tipo:
        qs = qs.filter(objeto_tipo=objeto_tipo)
    if permissao:
        qs = qs.filter(permissao__icontains=permissao)
    if busca:
        qs = qs.filter(
            Q(descricao__icontains=busca) |
            Q(objeto_tipo__icontains=busca) |
            Q(objeto_id__icontains=busca) |
            Q(username__icontains=busca) |
            Q(nome_guerra__icontains=busca)
        )
    if data_inicio:
        qs = qs.filter(criado_em__date__gte=data_inicio)
    if data_fim:
        qs = qs.filter(criado_em__date__lte=data_fim)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return JsonResponse({
        'results': [{
            'id': log.id,
            'linha': log.linha_formatada,
            'username': log.username,
            'nome_guerra': log.nome_guerra,
            'permissao': log.permissao,
            'secao': log.secao,
            'acao': log.acao,
            'objeto_tipo': log.objeto_tipo,
            'objeto_id': log.objeto_id,
            'descricao': log.descricao,
            'criado_em': log.criado_em.isoformat(),
        } for log in page_obj],
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'page': page_obj.number,
        'num_pages': page_obj.paginator.num_pages,
        'count': page_obj.paginator.count,
    })


@login_required
def auditoria_filtros(request):
    """Opções disponíveis pra montar os selects de filtro na página de Auditoria."""
    if not is_informatica_secao(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from auditoria.models import LogAuditoria
    secoes = list(LogAuditoria.objects.exclude(secao='').values_list('secao', flat=True).distinct().order_by('secao'))
    acoes = list(LogAuditoria.objects.exclude(acao='').values_list('acao', flat=True).distinct().order_by('acao'))
    objeto_tipos = list(LogAuditoria.objects.exclude(objeto_tipo='').values_list('objeto_tipo', flat=True).distinct().order_by('objeto_tipo'))
    permissoes = list(LogAuditoria.objects.exclude(permissao__in=['', '—']).values_list('permissao', flat=True).distinct().order_by('permissao'))
    return JsonResponse({'secoes': secoes, 'acoes': acoes, 'objeto_tipos': objeto_tipos, 'permissoes': permissoes})