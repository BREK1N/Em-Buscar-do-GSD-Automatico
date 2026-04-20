import io
import json
import logging
import os
from datetime import timedelta

from django.conf import settings

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import authenticate
from django.views.generic import ListView
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.core.files.base import ContentFile
from django.core.files import File
from django.db import transaction
from django.db.models import Q, Count
from django.db.models.functions import TruncMonth

from ..models import PATD, Configuracao, Anexo
from Secao_pessoal.models import Efetivo
from ..forms import ComandanteAprovarForm
from .decorators import (
    comandante_redirect, ouvidoria_required, comandante_required,
    oficial_responsavel_required, ComandanteAccessMixin,
)
from .helpers import _sync_oficial_signature, get_document_pages

logger = logging.getLogger(__name__)

def _check_and_advance_reconsideracao_status(patd_pk):
    """
    Função centralizada que verifica, dentro de uma transação para garantir a consistência dos dados,
    se a reconsideração tem conteúdo e assinatura. Se ambas as condições forem verdadeiras,
    o status é avançado.
    """
    try:
        with transaction.atomic():
            # Bloqueia a linha da PATD para evitar condições de corrida
            patd = PATD.objects.select_for_update().get(pk=patd_pk)

            # --- CORREÇÃO: Forçar a releitura do objeto do banco de dados ---
            # Garante que estamos a verificar o estado mais recente, incluindo
            # a atualização da assinatura que acabou de ser salva.
            patd.refresh_from_db()

            # Detailed logging for debugging
            logger.info(f"PATD {patd.pk} (Reconsideration Check):")
            logger.info(f"  - Current status: {patd.status}")
            logger.info(f"  - Has texto_reconsideracao: {bool(patd.texto_reconsideracao)}")
            logger.info(f"  - Anexos (tipo='reconsideracao') count: {patd.anexos.filter(tipo='reconsideracao').count()}")
            logger.info(f"  - Has assinatura_reconsideracao: {bool(patd.assinatura_reconsideracao)}")

            has_content = bool(patd.texto_reconsideracao or patd.anexos.filter(tipo='reconsideracao').exists())
            
            # Primeiro, tenta verificar pelo campo do modelo (que deveria estar atualizado)
            has_signature_db_field = bool(patd.assinatura_reconsideracao)
            logger.info(f"  - Has assinatura_reconsideracao (DB field): {has_signature_db_field}")

            # Se o campo do DB ainda for False, verifica diretamente no sistema de ficheiros como fallback
            has_signature_filesystem = False
            if not has_signature_db_field and patd.assinatura_reconsideracao and patd.assinatura_reconsideracao.name:
                try:
                    # Constrói o caminho absoluto esperado para o ficheiro da assinatura
                    expected_file_path = os.path.join(settings.MEDIA_ROOT, patd.assinatura_reconsideracao.name)
                    has_signature_filesystem = os.path.exists(expected_file_path)
                    logger.info(f"  - Has assinatura_reconsideracao (Filesystem check at {expected_file_path}): {has_signature_filesystem}")
                except Exception as e:
                    logger.warning(f"Erro ao verificar o sistema de ficheiros para a assinatura da PATD {patd.pk} no caminho '{patd.assinatura_reconsideracao.name}': {e}")

            # A assinatura é considerada válida se estiver no DB OU no sistema de ficheiros
            final_has_signature = has_signature_db_field or has_signature_filesystem

            if patd.status == 'em_reconsideracao' and has_content and final_has_signature:
                patd.status = 'aguardando_comandante_base'
                patd.save(update_fields=['status'])
                logger.info(f"PATD {patd.pk} status advanced to 'aguardando_comandante_base'.")
            else:
                logger.info(f"PATD {patd.pk}: Conditions not met to advance status (still 'em_reconsideracao').")
    except PATD.DoesNotExist:
        logger.error(f"PATD {patd_pk} not found during status check.")
    except Exception as e:
        logger.error(f"Error in _check_and_advance_reconsideracao_status for PATD {patd_pk}: {e}")


def _check_and_finalize_patd(patd):
 
    if patd.status != 'aguardando_assinatura_npd':
        return False

    document_pages = get_document_pages(patd)
    raw_document_text = "".join(document_pages)

    required_mil_signatures = raw_document_text.count('{Assinatura Militar Arrolado}')
    provided_mil_signatures = sum(1 for s in (patd.assinaturas_militar or []) if s)
    if provided_mil_signatures < required_mil_signatures:
        return False

    if not patd.testemunha1 or not patd.assinatura_testemunha1:
        return False

    if not patd.testemunha2 or not patd.assinatura_testemunha2:
        return False

    patd.status = 'periodo_reconsideracao'
    patd.data_publicacao_punicao = timezone.now()
    patd.save()
    return True


@method_decorator([login_required, comandante_required], name='dispatch')
class ComandanteDashboardView(ListView):
    model = PATD
    template_name = 'comandante_dashboard.html'
    context_object_name = 'patds'

    def get_queryset(self):
        # Este queryset é para a lista principal de PATDs "Aguardando Decisão"
        return PATD.objects.filter(status='analise_comandante').select_related('militar').order_by('-data_inicio')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        start_of_month = today.replace(day=1)

        # Métricas de contagem simples
        context['patd_em_andamento'] = PATD.objects.exclude(Q(status='finalizado') | Q(justificado=True)).count()
        context['patd_finalizadas_total'] = PATD.objects.filter(status='finalizado').count()
        context['patd_justificadas_total'] = PATD.objects.filter(justificado=True).count()
        
        # Criadas na semana/mês
        context['patd_criadas_semana'] = PATD.objects.filter(data_inicio__date__gte=start_of_week).count()
        context['patd_criadas_mes'] = PATD.objects.filter(data_inicio__date__gte=start_of_month).count()

        # Finalizadas na semana/mês
        context['patd_finalizadas_semana'] = PATD.objects.filter(status='finalizado', data_termino__date__gte=start_of_week).count()
        context['patd_finalizadas_mes'] = PATD.objects.filter(status='finalizado', data_termino__date__gte=start_of_month).count()

        # Dados para gráficos
        # Garante que todos os meses nos últimos 12 tenham um valor
        labels = []
        criadas_counts = []
        finalizadas_counts = []
        
        criadas_por_mes_dict = {
            item['month'].strftime('%Y-%m'): item['count']
            for item in PATD.objects
            .annotate(month=TruncMonth('data_inicio'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        }
        
        finalizadas_por_mes_dict = {
            item['month'].strftime('%Y-%m'): item['count']
            for item in PATD.objects
            .filter(status='finalizado')
            .annotate(month=TruncMonth('data_termino'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        }

        for i in range(11, -1, -1):
            current_month = today - timedelta(days=i*30)
            month_key = current_month.strftime('%Y-%m')
            labels.append(current_month.strftime('%b/%y'))
            
            criadas_counts.append(criadas_por_mes_dict.get(month_key, 0))
            finalizadas_counts.append(finalizadas_por_mes_dict.get(month_key, 0))

        context['chart_labels'] = json.dumps(labels)
        context['chart_data_criadas'] = json.dumps(criadas_counts)
        context['chart_data_finalizadas'] = json.dumps(finalizadas_counts)

        context['status_choices'] = PATD.STATUS_CHOICES
        context['oficiais_list'] = Efetivo.objects.filter(oficial=True).order_by('posto', 'nome_guerra')
        context['militares_list'] = Efetivo.objects.filter(deleted=False).order_by('posto', 'nome_guerra')

        return context


@login_required
@comandante_required
@require_POST # Garante que só aceita POST
def patd_aprovar(request, pk):
    patd = get_object_or_404(PATD, pk=pk)
    form = ComandanteAprovarForm(request.POST)

    # Verifica se as testemunhas estão definidas (lógica existente)
    errors = []
    if not patd.testemunha1 or not patd.testemunha2:
        errors.append("É necessário definir as duas testemunhas no processo.")

    if errors:
        error_message = f"PATD Nº {patd.numero_patd}: Não foi possível aprovar. " + " ".join(errors)
        messages.error(request, error_message)
        return redirect(request.META.get('HTTP_REFERER', 'Ouvidoria:comandante_dashboard'))

    # Verifica o formulário e a senha
    if form.is_valid():
        senha = form.cleaned_data['senha_comandante']
        # Autentica o usuário logado (que deve ser o comandante) com a senha fornecida
        user = authenticate(username=request.user.username, password=senha)

        if user is not None:
            # Senha correta, prossegue com a aprovação
            patd.status = 'aguardando_assinatura_npd'
            patd.save()
            messages.success(request, f"PATD Nº {patd.numero_patd} aprovada com sucesso. Aguardando assinatura da NPD.")
            return redirect('Ouvidoria:comandante_dashboard')
        else:
            # Senha incorreta
            messages.error(request, "Senha do Comandante incorreta. Aprovação não realizada.")
    else:
        # Formulário inválido (deve ter faltado a senha)
        messages.error(request, "Erro no formulário. A senha é obrigatória.")

    # Redireciona de volta em caso de erro de senha ou formulário
    return redirect(request.META.get('HTTP_REFERER', 'Ouvidoria:comandante_dashboard'))


@login_required
@comandante_required
@require_POST
def patd_retornar(request, pk):
    patd = get_object_or_404(PATD, pk=pk)
    comentario = request.POST.get('comentario')

    if not comentario:
        messages.error(request, "O comentário é obrigatório para retornar a PATD.")
        return redirect(request.META.get('HTTP_REFERER', 'Ouvidoria:comandante_dashboard'))

    patd.status = 'aguardando_punicao_alterar'
    patd.comentario_comandante = comentario
    patd.save()
    messages.warning(request, f"PATD Nº {patd.numero_patd} retornada para alteração com observações.")
    # Não precisa de senha aqui
    return redirect(request.META.get('HTTP_REFERER', 'Ouvidoria:comandante_dashboard'))


@login_required
@oficial_responsavel_required
@require_POST
def avancar_para_comandante(request, pk):
    patd = get_object_or_404(PATD, pk=pk)

    if not patd.testemunha1 or not patd.testemunha2:
        messages.error(request, "As testemunhas devem estar definidas antes de avançar para o Comandante.")
        return redirect('Ouvidoria:patd_detail', pk=pk)

    patd.status = 'analise_comandante'
    patd.save()
    messages.success(request, f"PATD Nº {patd.numero_patd} enviada para análise do Comandante.")
    return redirect('Ouvidoria:patd_detail', pk=pk)


@login_required
@ouvidoria_required
@require_POST
def solicitar_reconsideracao(request, pk):
    try:
        patd = get_object_or_404(PATD, pk=pk)
        if patd.status != 'periodo_reconsideracao':
            return JsonResponse({'status': 'error', 'message': 'A PATD não está no período de reconsideração.'}, status=400)

        patd.status = 'em_reconsideracao'
        patd.save(update_fields=['status'])
        messages.success(request, f'PATD Nº {patd.numero_patd} movida para "Em Reconsideração".')
        return JsonResponse({'status': 'success', 'message': 'Status atualizado com sucesso.'})
    except Exception as e:
        logger.error(f"Erro ao solicitar reconsideração para PATD {pk}: {e}")
        return JsonResponse({'status': 'error', 'message': 'Ocorreu um erro interno.'}, status=500)


@login_required
@ouvidoria_required
@require_POST
def salvar_reconsideracao(request, pk):
    try:
        patd = get_object_or_404(PATD, pk=pk)
        if patd.status != 'em_reconsideracao':
            return JsonResponse({'status': 'error', 'message': 'A PATD não está em fase de reconsideração.'}, status=400)

        texto = request.POST.get('texto_reconsideracao', '')
        arquivos = request.FILES.getlist('anexos_reconsideracao')

        if not texto and not arquivos:
            return JsonResponse({'status': 'error', 'message': 'É necessário fornecer um texto ou anexar pelo menos um ficheiro.'}, status=400)

        patd.texto_reconsideracao = texto
        if not patd.data_reconsideracao:
            patd.data_reconsideracao = timezone.now()
        patd.save(update_fields=['texto_reconsideracao', 'data_reconsideracao'])

        for arquivo in arquivos:
            Anexo.objects.create(patd=patd, arquivo=arquivo, tipo='reconsideracao')

        # --- CORREÇÃO: Usar transaction.on_commit para consistência ---
        # Garante que a verificação só ocorra após o salvamento do texto/anexos
        # ser confirmado na base de dados.
        transaction.on_commit(lambda: _check_and_advance_reconsideracao_status(pk))

        return JsonResponse({'status': 'success', 'message': 'Pedido de reconsideração e anexos salvos com sucesso.'})
    except Exception as e:
        logger.error(f"Erro ao salvar texto de reconsideração para PATD {pk}: {e}")
        return JsonResponse({'status': 'error', 'message': 'Ocorreu um erro interno.'}, status=500)


@login_required
@oficial_responsavel_required
@require_POST
def anexar_documento_reconsideracao_oficial(request, pk):
    try:
        patd = get_object_or_404(PATD, pk=pk)
        if patd.status != 'aguardando_comandante_base':
            messages.error(request, "Ação não permitida no status atual.")
            return redirect('Ouvidoria:patd_detail', pk=pk)

        anexo_file = request.FILES.get('anexo_oficial')
        if not anexo_file:
            messages.error(request, "Nenhum ficheiro foi enviado.")
            return redirect('Ouvidoria:patd_detail', pk=pk)

        Anexo.objects.create(patd=patd, arquivo=anexo_file, tipo='reconsideracao_oficial')

        # --- INÍCIO DA MODIFICAÇÃO ---
        # Atualiza o status para aguardar a definição da nova punição
        patd.status = 'aguardando_nova_punicao'
        patd.save(update_fields=['status'])

        messages.success(request, "Documento anexado com sucesso! O processo agora aguarda a definição da nova punição.")
        return redirect('Ouvidoria:patd_detail', pk=pk)
        # --- FIM DA MODIFICAÇÃO ---
    except Exception as e:
        logger.error(f"Erro ao anexar documento de reconsideração oficial para PATD {pk}: {e}")
        messages.error(request, f"Ocorreu um erro ao anexar o documento: {e}")
        return redirect('Ouvidoria:patd_detail', pk=pk)


@login_required
@comandante_required
def relatorio_json(request):
    qs = PATD.objects.select_related('militar', 'oficial_responsavel').order_by('-data_inicio')

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    oficial_pk = request.GET.get('oficial')
    militar_pk = request.GET.get('militar')

    if data_inicio:
        qs = qs.filter(data_inicio__date__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_inicio__date__lte=data_fim)
    if status:
        qs = qs.filter(status=status)
    if oficial_pk:
        qs = qs.filter(oficial_responsavel__pk=oficial_pk)
    if militar_pk:
        qs = qs.filter(militar__pk=militar_pk)

    patds_data = []
    for p in qs[:200]:
        patds_data.append({
            'pk': p.pk,
            'numero': p.numero_patd or '—',
            'militar': str(p.militar) if p.militar else '—',
            'oficial': str(p.oficial_responsavel) if p.oficial_responsavel else '—',
            'data_inicio': p.data_inicio.strftime('%d/%m/%Y') if p.data_inicio else '—',
            'data_ocorrencia': p.data_ocorrencia.strftime('%d/%m/%Y') if p.data_ocorrencia else '—',
            'status_display': p.get_status_display(),
            'status': p.status,
        })

    total = qs.count()
    finalizadas = qs.filter(status='finalizado').count()
    aguardando = qs.filter(status='analise_comandante').count()
    em_andamento = total - finalizadas

    return JsonResponse({
        'patds': patds_data,
        'kpis': {
            'total': total,
            'finalizadas': finalizadas,
            'em_andamento': em_andamento,
            'aguardando_decisao': aguardando,
        }
    })


@login_required
@comandante_required
def relatorio_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference, Series
    from openpyxl.chart.label import DataLabelList

    qs = PATD.objects.select_related('militar', 'oficial_responsavel').order_by('-data_inicio')

    data_inicio = request.GET.get('data_inicio')
    data_fim    = request.GET.get('data_fim')
    status      = request.GET.get('status')
    oficial_pk  = request.GET.get('oficial')
    militar_pk  = request.GET.get('militar')

    if data_inicio: qs = qs.filter(data_inicio__date__gte=data_inicio)
    if data_fim:    qs = qs.filter(data_inicio__date__lte=data_fim)
    if status:      qs = qs.filter(status=status)
    if oficial_pk:  qs = qs.filter(oficial_responsavel__pk=oficial_pk)
    if militar_pk:  qs = qs.filter(militar__pk=militar_pk)

    # Estatísticas
    total           = qs.count()
    n_finalizadas   = qs.filter(status='finalizado').count()
    n_aguardando    = qs.filter(status='analise_comandante').count()
    n_em_andamento  = total - n_finalizadas

    wb = openpyxl.Workbook()

    # ── Estilos compartilhados ──────────────────────────────────────────
    def make_fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def make_border(color='D1D5DB'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    fill_dark    = make_fill('1E293B')
    fill_amber   = make_fill('F59E0B')
    fill_green   = make_fill('22C55E')
    fill_red     = make_fill('EF4444')
    fill_blue    = make_fill('3B82F6')
    fill_alt     = make_fill('F8FAFC')
    fill_kpi_bg  = make_fill('F1F5F9')
    border       = make_border()
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_mid     = Alignment(horizontal='left',   vertical='center')

    # ══════════════════════════════════════════════════════════════════════
    # ABA 1 — RESUMO + GRÁFICO
    # ══════════════════════════════════════════════════════════════════════
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    # Título
    ws_resumo.merge_cells('A1:F1')
    ws_resumo['A1'] = 'RELATÓRIO DE PROCESSOS PATD'
    ws_resumo['A1'].font = Font(bold=True, color='FFFFFF', size=16)
    ws_resumo['A1'].fill = fill_amber
    ws_resumo['A1'].alignment = center
    ws_resumo.row_dimensions[1].height = 36

    # Filtros
    filtros = []
    if data_inicio: filtros.append(f'De: {data_inicio}')
    if data_fim:    filtros.append(f'Até: {data_fim}')
    if status:      filtros.append(f'Status: {dict(PATD.STATUS_CHOICES).get(status, status)}')
    ws_resumo.merge_cells('A2:F2')
    ws_resumo['A2'] = ('Filtros aplicados: ' + ' | '.join(filtros)) if filtros else 'Filtros: Todos os registros'
    ws_resumo['A2'].font = Font(italic=True, color='64748B', size=9)
    ws_resumo['A2'].alignment = left_mid
    ws_resumo.row_dimensions[2].height = 16

    # Gerado em
    from django.utils import timezone as tz
    ws_resumo.merge_cells('A3:F3')
    ws_resumo['A3'] = f'Gerado em: {tz.now().strftime("%d/%m/%Y %H:%M")}'
    ws_resumo['A3'].font = Font(italic=True, color='94A3B8', size=9)
    ws_resumo['A3'].alignment = left_mid
    ws_resumo.row_dimensions[3].height = 14

    ws_resumo.row_dimensions[4].height = 10  # espaço

    # KPIs — cabeçalhos
    kpi_headers = ['Total', 'Em Andamento', 'Finalizadas', 'Aguardando Decisão']
    kpi_values  = [total, n_em_andamento, n_finalizadas, n_aguardando]
    kpi_fills   = [fill_dark, fill_blue, fill_green, fill_red]

    for col, (hdr, val, fill) in enumerate(zip(kpi_headers, kpi_values, kpi_fills), start=1):
        # Header
        h = ws_resumo.cell(row=5, column=col, value=hdr)
        h.font = Font(bold=True, color='FFFFFF', size=10)
        h.fill = fill
        h.alignment = center
        h.border = border
        ws_resumo.row_dimensions[5].height = 20
        # Valor
        v = ws_resumo.cell(row=6, column=col, value=val)
        v.font = Font(bold=True, size=22)
        v.fill = fill_kpi_bg
        v.alignment = center
        v.border = border
        ws_resumo.row_dimensions[6].height = 40

    # Dados ocultos para o gráfico (linha 9-12)
    ws_resumo['A9'] = 'Situação'
    ws_resumo['B9'] = 'Quantidade'
    chart_labels = ['Em Andamento', 'Finalizadas', 'Aguardando Decisão']
    chart_values = [n_em_andamento, n_finalizadas, n_aguardando]
    for i, (lbl, val) in enumerate(zip(chart_labels, chart_values)):
        ws_resumo.cell(row=10 + i, column=1, value=lbl)
        ws_resumo.cell(row=10 + i, column=2, value=val)

    # Gráfico de pizza
    pie = PieChart()
    pie.title = 'Distribuição de Processos PATD'
    pie.style = 10
    pie.width  = 18
    pie.height = 12

    data_ref   = Reference(ws_resumo, min_col=2, min_row=9, max_row=12)
    labels_ref = Reference(ws_resumo, min_col=1, min_row=10, max_row=12)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False
    pie.series[0].dLbls = pie.dataLabels

    ws_resumo.add_chart(pie, 'A8')

    # Gráfico de barras (ao lado)
    bar = BarChart()
    bar.type    = 'col'
    bar.title   = 'Quantidade por Situação'
    bar.style   = 10
    bar.width   = 18
    bar.height  = 12
    bar.y_axis.title = 'Processos'
    bar.x_axis.title = 'Situação'

    bar_data   = Reference(ws_resumo, min_col=2, min_row=9, max_row=12)
    bar_labels = Reference(ws_resumo, min_col=1, min_row=10, max_row=12)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_labels)
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True

    ws_resumo.add_chart(bar, 'D8')

    # Larguras colunas resumo
    for col, w in zip('ABCDEF', [20, 20, 20, 22, 10, 10]):
        ws_resumo.column_dimensions[col].width = w

    # Ocultar linhas de dados do gráfico
    for r in range(9, 13):
        ws_resumo.row_dimensions[r].hidden = True

    # ══════════════════════════════════════════════════════════════════════
    # ABA 2 — PROCESSOS (tabela completa)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Processos')

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'LISTA DE PROCESSOS'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = fill_dark
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    # Cabeçalho tabela
    headers = ['N° PATD', 'Militar Acusado', 'Posto', 'Oficial Apurador', 'Data Início', 'Data Ocorrência', 'Status', 'Punição', 'Itens Enquadrados']
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = fill_dark
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 22

    # Dados
    patds_list = list(qs)
    for row_idx, patd in enumerate(patds_list, start=3):
        is_alt   = (row_idx % 2 == 0)
        row_fill = fill_alt if is_alt else None
        itens_str = ', '.join([str(i.get('numero', '')) for i in (patd.itens_enquadrados or []) if i.get('numero')]) or '—'
        row_data = [
            patd.numero_patd or '—',
            patd.militar.nome_guerra if patd.militar else '—',
            patd.militar.posto if patd.militar else '—',
            str(patd.oficial_responsavel) if patd.oficial_responsavel else '—',
            patd.data_inicio.strftime('%d/%m/%Y') if patd.data_inicio else '—',
            patd.data_ocorrencia.strftime('%d/%m/%Y') if patd.data_ocorrencia else '—',
            patd.get_status_display(),
            getattr(patd, 'punicao', '') or '—',
            itens_str,
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = left_mid
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 18

    # Total
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=total_row, column=1).fill = fill_amber
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=total_row, column=2).fill = fill_amber
    ws.cell(row=total_row, column=2).alignment = center

    # Larguras colunas processos
    for col, w in enumerate([10, 22, 14, 24, 14, 16, 30, 20, 22], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Auto-filter e freeze
    ws.auto_filter.ref = f'A2:I{ws.max_row}'
    ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════════════════════
    # ABA 3 — POR OFICIAL
    # ══════════════════════════════════════════════════════════════════════
    from collections import defaultdict
    ws_of = wb.create_sheet('Por Oficial')

    ws_of.merge_cells('A1:F1')
    ws_of['A1'] = 'ANÁLISE POR OFICIAL APURADOR'
    ws_of['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws_of['A1'].fill = fill_dark
    ws_of['A1'].alignment = center
    ws_of.row_dimensions[1].height = 28

    hdr_of = ['Oficial Apurador', 'Total', 'Finalizadas', 'Em Andamento', 'Aguard. Decisão', '% Finalizado']
    for col, h in enumerate(hdr_of, start=1):
        cell = ws_of.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = fill_dark
        cell.alignment = center
        cell.border = border
    ws_of.row_dimensions[2].height = 22

    stats_of = defaultdict(lambda: {'total': 0, 'fin': 0, 'agu': 0})
    for patd in patds_list:
        nome = str(patd.oficial_responsavel) if patd.oficial_responsavel else '(sem oficial)'
        stats_of[nome]['total'] += 1
        if patd.status == 'finalizado':
            stats_of[nome]['fin'] += 1
        elif patd.status == 'analise_comandante':
            stats_of[nome]['agu'] += 1

    for row_idx, (nome, s) in enumerate(sorted(stats_of.items()), start=3):
        is_alt   = (row_idx % 2 == 0)
        row_fill = fill_alt if is_alt else None
        andamento = s['total'] - s['fin']
        pct = round(s['fin'] / s['total'] * 100, 1) if s['total'] else 0
        row_data = [nome, s['total'], s['fin'], andamento, s['agu'], f"{pct}%"]
        for col, value in enumerate(row_data, start=1):
            cell = ws_of.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = center if col > 1 else left_mid
            if row_fill:
                cell.fill = row_fill
        ws_of.row_dimensions[row_idx].height = 18

    for col, w in enumerate([30, 10, 14, 16, 18, 14], start=1):
        ws_of.column_dimensions[get_column_letter(col)].width = w

    ws_of.auto_filter.ref = f'A2:F{ws_of.max_row}'
    ws_of.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════════════════════
    # ABA 4 — POR MILITAR
    # ══════════════════════════════════════════════════════════════════════
    ws_mil = wb.create_sheet('Por Militar')

    ws_mil.merge_cells('A1:F1')
    ws_mil['A1'] = 'ANÁLISE POR MILITAR ACUSADO'
    ws_mil['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws_mil['A1'].fill = fill_dark
    ws_mil['A1'].alignment = center
    ws_mil.row_dimensions[1].height = 28

    hdr_mil = ['Militar', 'Posto', 'Total PATDs', 'Finalizadas', 'Em Andamento', 'Itens Mais Frequentes']
    for col, h in enumerate(hdr_mil, start=1):
        cell = ws_mil.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = fill_dark
        cell.alignment = center
        cell.border = border
    ws_mil.row_dimensions[2].height = 22

    stats_mil = defaultdict(lambda: {'posto': '', 'total': 0, 'fin': 0, 'itens': defaultdict(int)})
    for patd in patds_list:
        if not patd.militar:
            continue
        key = patd.militar.nome_guerra
        stats_mil[key]['posto'] = patd.militar.posto or ''
        stats_mil[key]['total'] += 1
        if patd.status == 'finalizado':
            stats_mil[key]['fin'] += 1
        for item in (patd.itens_enquadrados or []):
            n = item.get('numero')
            if n:
                stats_mil[key]['itens'][str(n)] += 1

    sorted_mil = sorted(stats_mil.items(), key=lambda x: x[1]['total'], reverse=True)
    for row_idx, (nome, s) in enumerate(sorted_mil, start=3):
        is_alt   = (row_idx % 2 == 0)
        row_fill = fill_alt if is_alt else None
        top_itens = ', '.join([f"Item {k}({v}x)" for k, v in sorted(s['itens'].items(), key=lambda x: -x[1])[:5]]) or '—'
        row_data = [nome, s['posto'], s['total'], s['fin'], s['total'] - s['fin'], top_itens]
        for col, value in enumerate(row_data, start=1):
            cell = ws_mil.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = center if col in (3, 4, 5) else left_mid
            if row_fill:
                cell.fill = row_fill
        ws_mil.row_dimensions[row_idx].height = 18

    for col, w in enumerate([24, 14, 12, 12, 14, 32], start=1):
        ws_mil.column_dimensions[get_column_letter(col)].width = w

    ws_mil.auto_filter.ref = f'A2:F{ws_mil.max_row}'
    ws_mil.freeze_panes = 'A3'

    # Salvar e retornar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'relatorio_patd_{tz.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
