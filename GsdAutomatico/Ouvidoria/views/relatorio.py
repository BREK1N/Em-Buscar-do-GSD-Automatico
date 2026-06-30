import io
import logging
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone as tz
from django.views.decorators.http import require_GET

from ..models import PATD
from ..permissions import has_ouvidoria_access
from .decorators import ouvidoria_required
from Secao_pessoal.models import Efetivo

logger = logging.getLogger(__name__)

# Grupos de status para coloração
STATUS_NAO_INICIADO = {'definicao_oficial', 'aguardando_aprovacao_atribuicao'}
STATUS_FINALIZADO   = {'finalizado'}


def _cor_status(status):
    if status in STATUS_NAO_INICIADO:
        return 'nao_iniciado'
    if status in STATUS_FINALIZADO:
        return 'finalizado'
    return 'em_andamento'


def status_label(patd):
    """Rótulo de status exibido nos relatórios — sinaliza PATDs arquivadas."""
    if patd.arquivado:
        return 'Arquivada'
    return patd.get_status_display()


def origem_label(patd):
    return 'Sistema Antigo' if patd.sistema_antigo else 'Atual'


@login_required
@ouvidoria_required
def relatorio_ouvidoria(request):
    return render(request, 'relatorio_ouvidoria.html', {
        'status_choices':  PATD.STATUS_CHOICES,
        'oficiais_list':   Efetivo.objects.filter(oficial=True).order_by('posto', 'nome_guerra'),
        'militares_list':  Efetivo.objects.filter(deleted=False).order_by('posto', 'nome_guerra'),
    })


@login_required
@ouvidoria_required
@require_GET
def relatorio_ouvidoria_json(request):
    qs = (
        PATD.objects
        .select_related('militar', 'oficial_responsavel')
        .filter(deleted=False)
        .order_by('-data_inicio')
    )

    data_inicio = request.GET.get('data_inicio')
    data_fim    = request.GET.get('data_fim')
    status      = request.GET.get('status')
    oficial_pk  = request.GET.get('oficial')
    militar_pk  = request.GET.get('militar')
    setor       = request.GET.get('setor')

    if data_inicio: qs = qs.filter(data_inicio__date__gte=data_inicio)
    if data_fim:    qs = qs.filter(data_inicio__date__lte=data_fim)
    if status:      qs = qs.filter(status=status)
    if oficial_pk:  qs = qs.filter(oficial_responsavel__pk=oficial_pk)
    if militar_pk:  qs = qs.filter(militar__pk=militar_pk)
    if setor:       qs = qs.filter(militar__setor__icontains=setor)

    total         = qs.count()
    n_finalizado  = qs.filter(status='finalizado').count()
    n_nao_inic    = qs.filter(status__in=STATUS_NAO_INICIADO).count()
    n_andamento   = total - n_finalizado - n_nao_inic

    patds_data = []
    for p in qs[:500]:
        itens_str = ', '.join(
            [str(i.get('numero', '')) for i in (p.itens_enquadrados or []) if i.get('numero')]
        ) or '—'
        patds_data.append({
            'pk': p.pk,
            'numero':           p.numero_patd or (f'(Antigo) {p.numero_patd_legado}' if p.numero_patd_legado else '—'),
            'saram':            str(p.militar.saram) if p.militar and p.militar.saram else '—',
            'militar':          p.militar.nome_guerra if p.militar else '—',
            'nome_completo':    p.militar.nome_completo if p.militar else '—',
            'posto':            p.militar.posto if p.militar else '—',
            'setor':            p.militar.setor if p.militar else '—',
            'oficial':          str(p.oficial_responsavel) if p.oficial_responsavel else '—',
            'data_inicio':      p.data_inicio.strftime('%d/%m/%Y') if p.data_inicio else '—',
            'data_ocorrencia':  p.data_inicio.strftime('%d/%m/%Y') if p.data_inicio else '—',
            'protocolo_comaer': p.protocolo_comaer or '—',
            'status_display':   status_label(p),
            'status':           p.status,
            'cor':              'arquivada' if p.arquivado else _cor_status(p.status),
            'origem':           origem_label(p),
            'natureza':         p.natureza_transgressao or '—',
            'punicao':          p.punicao or '—',
            'dias_punicao':     p.dias_punicao or '—',
            'itens':            itens_str,
            'transgressao':     (p.transgressao or '')[:200],
            'texto_relatorio':  (p.texto_relatorio or '')[:300],
        })

    return JsonResponse({
        'patds': patds_data,
        'kpis': {
            'total':       total,
            'finalizado':  n_finalizado,
            'em_andamento': n_andamento,
            'nao_iniciado': n_nao_inic,
        }
    })


@login_required
@ouvidoria_required
@require_GET
def relatorio_ouvidoria_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    qs = (
        PATD.objects
        .select_related('militar', 'oficial_responsavel')
        .filter(deleted=False)
        .order_by('-data_inicio')
    )

    data_inicio = request.GET.get('data_inicio')
    data_fim    = request.GET.get('data_fim')
    status      = request.GET.get('status')
    oficial_pk  = request.GET.get('oficial')
    militar_pk  = request.GET.get('militar')
    setor       = request.GET.get('setor')

    if data_inicio: qs = qs.filter(data_inicio__date__gte=data_inicio)
    if data_fim:    qs = qs.filter(data_inicio__date__lte=data_fim)
    if status:      qs = qs.filter(status=status)
    if oficial_pk:  qs = qs.filter(oficial_responsavel__pk=oficial_pk)
    if militar_pk:  qs = qs.filter(militar__pk=militar_pk)
    if setor:       qs = qs.filter(militar__setor__icontains=setor)

    patds_list   = list(qs)
    total        = len(patds_list)
    n_finalizado = sum(1 for p in patds_list if p.status == 'finalizado')
    n_nao_inic   = sum(1 for p in patds_list if p.status in STATUS_NAO_INICIADO)
    n_andamento  = total - n_finalizado - n_nao_inic

    # ── Helpers de estilo ──────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def border(color='D1D5DB'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    F_DARK    = fill('1E293B')
    F_GREEN   = fill('22C55E')   # finalizado
    F_YELLOW  = fill('FDE68A')   # em andamento
    F_WHITE   = fill('FFFFFF')   # não iniciado
    F_AMBER   = fill('F59E0B')
    F_BLUE    = fill('3B82F6')
    F_ALT     = fill('F8FAFC')
    F_KPI     = fill('F1F5F9')
    BRD       = border()

    def status_fill(status):
        if status in STATUS_FINALIZADO:
            return F_GREEN
        if status in STATUS_NAO_INICIADO:
            return F_WHITE
        return F_YELLOW

    def status_font_color(status):
        if status in STATUS_FINALIZADO:
            return '14532D'
        if status in STATUS_NAO_INICIADO:
            return '374151'
        return '78350F'

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # ABA 1 — RESUMO
    # ══════════════════════════════════════════════════════════════════════
    ws_r = wb.active
    ws_r.title = 'Resumo'

    ws_r.merge_cells('A1:G1')
    ws_r['A1'] = 'RELATÓRIO DA OUVIDORIA — PROCESSOS PATD'
    ws_r['A1'].font      = Font(bold=True, color='FFFFFF', size=16)
    ws_r['A1'].fill      = F_AMBER
    ws_r['A1'].alignment = CENTER
    ws_r.row_dimensions[1].height = 36

    filtros = []
    if data_inicio: filtros.append(f'De: {data_inicio}')
    if data_fim:    filtros.append(f'Até: {data_fim}')
    if status:      filtros.append(f'Status: {dict(PATD.STATUS_CHOICES).get(status, status)}')
    if oficial_pk:
        try:
            of = Efetivo.objects.get(pk=oficial_pk)
            filtros.append(f'Oficial: {of.nome_guerra}')
        except Efetivo.DoesNotExist:
            pass
    if setor:       filtros.append(f'Setor: {setor}')
    ws_r.merge_cells('A2:G2')
    ws_r['A2'] = ('Filtros: ' + ' | '.join(filtros)) if filtros else 'Filtros: Todos os registros'
    ws_r['A2'].font = Font(italic=True, color='64748B', size=9)
    ws_r['A2'].alignment = LEFT
    ws_r.row_dimensions[2].height = 14

    ws_r.merge_cells('A3:G3')
    ws_r['A3'] = f'Gerado em: {tz.now().strftime("%d/%m/%Y %H:%M")}'
    ws_r['A3'].font = Font(italic=True, color='94A3B8', size=9)
    ws_r['A3'].alignment = LEFT
    ws_r.row_dimensions[3].height = 13

    ws_r.row_dimensions[4].height = 8

    # KPIs
    kpi_data = [
        ('Total',          total,        F_DARK,  'FFFFFF'),
        ('Não Iniciado',   n_nao_inic,   F_BLUE,  'FFFFFF'),
        ('Em Andamento',   n_andamento,  F_AMBER, 'FFFFFF'),
        ('Finalizados',    n_finalizado, F_GREEN, 'FFFFFF'),
    ]
    for col, (lbl, val, bg, fc) in enumerate(kpi_data, start=1):
        h = ws_r.cell(row=5, column=col, value=lbl)
        h.font = Font(bold=True, color=fc, size=10); h.fill = bg; h.alignment = CENTER; h.border = BRD
        v = ws_r.cell(row=6, column=col, value=val)
        v.font = Font(bold=True, size=22); v.fill = F_KPI; v.alignment = CENTER; v.border = BRD
    ws_r.row_dimensions[5].height = 20
    ws_r.row_dimensions[6].height = 42

    # Legenda de cores
    ws_r.row_dimensions[7].height = 8
    ws_r['A8'] = 'LEGENDA DE CORES NA ABA "PROCESSOS":'
    ws_r['A8'].font = Font(bold=True, size=9, color='374151')
    ws_r.merge_cells('A8:G8')

    leg = [
        ('B9',  'Não Iniciado',  F_WHITE,  '374151'),
        ('C9',  'Em Andamento',  F_YELLOW, '78350F'),
        ('D9',  'Finalizado',    F_GREEN,  '14532D'),
    ]
    for cell_ref, lbl, bg, fc in leg:
        c = ws_r[cell_ref]
        c.value = lbl; c.fill = bg
        c.font = Font(bold=True, color=fc, size=9)
        c.alignment = CENTER; c.border = BRD
    ws_r.row_dimensions[9].height = 18

    # Dados gráfico (ocultos)
    for r, (lbl, val) in enumerate(zip(['Não Iniciado', 'Em Andamento', 'Finalizado'],
                                        [n_nao_inic, n_andamento, n_finalizado]), start=12):
        ws_r.cell(row=r, column=1, value=lbl)
        ws_r.cell(row=r, column=2, value=val)

    pie = PieChart()
    pie.title = 'Distribuição por Situação'
    pie.style = 10; pie.width = 18; pie.height = 12
    pie.add_data(Reference(ws_r, min_col=2, min_row=11, max_row=14), titles_from_data=True)
    pie.set_categories(Reference(ws_r, min_col=1, min_row=12, max_row=14))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    ws_r.add_chart(pie, 'A11')

    bar = BarChart()
    bar.type = 'col'; bar.title = 'Quantidade por Situação'
    bar.style = 10; bar.width = 18; bar.height = 12
    bar.add_data(Reference(ws_r, min_col=2, min_row=11, max_row=14), titles_from_data=True)
    bar.set_categories(Reference(ws_r, min_col=1, min_row=12, max_row=14))
    bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
    ws_r.add_chart(bar, 'D11')

    for r in range(11, 15):
        ws_r.row_dimensions[r].hidden = True

    for col, w in zip('ABCDEFG', [22, 22, 22, 22, 12, 12, 12]):
        ws_r.column_dimensions[col].width = w

    # ══════════════════════════════════════════════════════════════════════
    # ABA 2 — PROCESSOS (tabela completa com cores)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Processos')

    ws.merge_cells('A1:M1')
    ws['A1'] = 'LISTA DE PROCESSOS PATD'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = F_DARK; ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = [
        'N° PATD', 'SARAM', 'Militar', 'Nome Completo', 'Posto',
        'Setor', 'Oficial Apurador', 'Data de Início',
        'Status', 'Natureza', 'Punição', 'Itens Enquadrados', 'Origem',
    ]
    for col, hdr in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = F_DARK; c.alignment = CENTER; c.border = BRD
    ws.row_dimensions[2].height = 22

    for row_idx, patd in enumerate(patds_list, start=3):
        sf   = status_fill(patd.status)
        sfc  = status_font_color(patd.status)
        alt  = F_ALT if row_idx % 2 == 0 else None
        itens_str = ', '.join(
            [str(i.get('numero', '')) for i in (patd.itens_enquadrados or []) if i.get('numero')]
        ) or '—'

        row_data = [
            patd.numero_patd or (f'(Antigo) {patd.numero_patd_legado}' if patd.numero_patd_legado else '—'),
            str(patd.militar.saram) if patd.militar and patd.militar.saram else '—',
            patd.militar.nome_guerra if patd.militar else '—',
            patd.militar.nome_completo if patd.militar else '—',
            patd.militar.posto if patd.militar else '—',
            patd.militar.setor if patd.militar else '—',
            str(patd.oficial_responsavel) if patd.oficial_responsavel else '—',
            patd.data_inicio.strftime('%d/%m/%Y') if patd.data_inicio else '—',
            status_label(patd),
            patd.natureza_transgressao or '—',
            f"{patd.dias_punicao or ''} {patd.punicao or ''}".strip() or '—',
            itens_str,
            origem_label(patd),
        ]
        for col, value in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.border = BRD
            c.alignment = LEFT
            # Coluna de status recebe a cor temática; demais recebem alternância
            if col == 9:
                c.fill = sf
                c.font = Font(bold=True, color=sfc, size=9)
            elif alt:
                c.fill = alt
        ws.row_dimensions[row_idx].height = 18

    # Rodapé total
    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=tr, column=1).fill = F_AMBER; ws.cell(row=tr, column=1).alignment = CENTER
    ws.cell(row=tr, column=2, value=total).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=tr, column=2).fill = F_AMBER; ws.cell(row=tr, column=2).alignment = CENTER

    for col, w in enumerate([10, 10, 18, 28, 10, 16, 24, 14, 32, 16, 20, 22, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{ws.max_row}'
    ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════════════════════
    # ABA 3 — RESUMO DA APURAÇÃO (transgressão + texto do relatório)
    # ══════════════════════════════════════════════════════════════════════
    ws_ap = wb.create_sheet('Resumo Apuração')

    ws_ap.merge_cells('A1:E1')
    ws_ap['A1'] = 'RESUMO DAS APURAÇÕES'
    ws_ap['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws_ap['A1'].fill = F_DARK; ws_ap['A1'].alignment = CENTER
    ws_ap.row_dimensions[1].height = 28

    hdr_ap = ['N° PATD', 'Militar', 'Transgressão', 'Resumo do Relatório', 'Status']
    for col, h in enumerate(hdr_ap, start=1):
        c = ws_ap.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = F_DARK; c.alignment = CENTER; c.border = BRD
    ws_ap.row_dimensions[2].height = 22

    # Maior N° PATD primeiro (mais recente -> mais antigo); sem nº fica por último
    patds_apuracao = sorted(
        patds_list, key=lambda p: (p.numero_patd is None, -(p.numero_patd or 0))
    )

    for row_idx, patd in enumerate(patds_apuracao, start=3):
        sf  = status_fill(patd.status)
        sfc = status_font_color(patd.status)
        alt = F_ALT if row_idx % 2 == 0 else None
        row_data = [
            patd.numero_patd or (f'(Antigo) {patd.numero_patd_legado}' if patd.numero_patd_legado else '—'),
            f"{patd.militar.posto or ''} {patd.militar.nome_guerra or ''}".strip() if patd.militar else '—',
            patd.transgressao or '—',
            patd.texto_relatorio or '—',
            status_label(patd),
        ]
        for col, value in enumerate(row_data, start=1):
            c = ws_ap.cell(row=row_idx, column=col, value=str(value))
            c.border = BRD
            c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if col == 5:
                c.fill = sf; c.font = Font(bold=True, color=sfc, size=9)
            elif alt:
                c.fill = alt
        ws_ap.row_dimensions[row_idx].height = 60

    for col, w in enumerate([10, 22, 60, 70, 30], start=1):
        ws_ap.column_dimensions[get_column_letter(col)].width = w

    ws_ap.auto_filter.ref = f'A2:E{ws_ap.max_row}'
    ws_ap.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════════════════════
    # ABA 4 — POR SETOR
    # ══════════════════════════════════════════════════════════════════════
    ws_set = wb.create_sheet('Por Setor')

    ws_set.merge_cells('A1:F1')
    ws_set['A1'] = 'ANÁLISE POR SETOR'
    ws_set['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws_set['A1'].fill = F_DARK; ws_set['A1'].alignment = CENTER
    ws_set.row_dimensions[1].height = 28

    hdr_set = ['Setor', 'Total', 'Não Iniciado', 'Em Andamento', 'Finalizado', '% Finalizado']
    for col, h in enumerate(hdr_set, start=1):
        c = ws_set.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = F_DARK; c.alignment = CENTER; c.border = BRD
    ws_set.row_dimensions[2].height = 22

    stats_set = defaultdict(lambda: {'total': 0, 'fin': 0, 'ni': 0})
    for patd in patds_list:
        s_key = (patd.militar.setor or '(sem setor)') if patd.militar else '(sem setor)'
        stats_set[s_key]['total'] += 1
        if patd.status in STATUS_FINALIZADO:
            stats_set[s_key]['fin'] += 1
        elif patd.status in STATUS_NAO_INICIADO:
            stats_set[s_key]['ni'] += 1

    for row_idx, (setor_nome, s) in enumerate(
            sorted(stats_set.items(), key=lambda x: x[1]['total'], reverse=True), start=3):
        alt = F_ALT if row_idx % 2 == 0 else None
        and_ = s['total'] - s['fin'] - s['ni']
        pct  = round(s['fin'] / s['total'] * 100, 1) if s['total'] else 0
        row_data = [setor_nome, s['total'], s['ni'], and_, s['fin'], f"{pct}%"]
        for col, value in enumerate(row_data, start=1):
            c = ws_set.cell(row=row_idx, column=col, value=value)
            c.border = BRD
            c.alignment = CENTER if col > 1 else LEFT
            if alt: c.fill = alt
        ws_set.row_dimensions[row_idx].height = 18

    for col, w in enumerate([28, 10, 14, 16, 14, 14], start=1):
        ws_set.column_dimensions[get_column_letter(col)].width = w

    ws_set.auto_filter.ref = f'A2:F{ws_set.max_row}'
    ws_set.freeze_panes = 'A3'

    # ── Salvar e devolver ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'relatorio_ouvidoria_{tz.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
